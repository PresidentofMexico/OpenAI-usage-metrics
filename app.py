"""
Multi-Tool AI Usage Analytics Dashboard

A Streamlit-based dashboard for analyzing AI tool usage metrics across multiple platforms
(OpenAI ChatGPT, BlueFlame AI, etc.) for enterprise management reporting.
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import glob
from io import StringIO
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import sqlite3
import json
import traceback

from data_processor import DataProcessor
from database import DatabaseManager
from file_reader import read_file_robust, display_file_error, read_file_from_path
from file_scanner import FileScanner
from config import AUTO_SCAN_FOLDERS, FILE_TRACKING_PATH, ENTERPRISE_PRICING, RECURSIVE_SCAN_FOLDERS
from export_utils import generate_excel_export, generate_pdf_report_html
from cost_calculator import EnterpriseCostCalculator

# Constants
WEEKLY_CHART_DATE_FORMAT = '%m/%d/%Y'  # Format for displaying week dates in weekly trend charts

# ROI Analytics Configuration
# These values can be customized based on organization-specific benchmarks
ROI_HOURLY_VALUE = 50  # Dollar value per hour of time saved (adjust based on average salary)
ROI_MONTHLY_WORK_HOURS = 160  # Standard monthly work hours (20 days × 8 hours/day)

# Page configuration
st.set_page_config(
    page_title="AI Usage Analytics Dashboard",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize database and data processor
@st.cache_resource
def init_app():
    db = DatabaseManager()
    processor = DataProcessor(db)
    scanner = FileScanner(FILE_TRACKING_PATH, recursive_folders=RECURSIVE_SCAN_FOLDERS)
    
    # Auto-load employee file if it exists
    auto_load_employee_file(db)
    
    return db, processor, scanner

def auto_load_employee_file(db_manager):
    """
    Automatically load employee master file from repository if it exists.
    This function uses glob patterns to detect employee CSV files and loads them automatically.
    
    Supported file patterns:
    - Employee Headcount*Emails.csv (preferred)
    - Employee Headcount*.csv (fallback)
    """
    # Get the directory where this script is located (not the current working directory)
    # This ensures we look for the employee file in the repository root, regardless of
    # where the app is launched from (important for cloud deployments)
    script_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"[auto_load_employee_file] Script directory: {script_dir}")
    print(f"[auto_load_employee_file] Current working directory: {os.getcwd()}")
    
    # Check if employees are already loaded - if so, we don't need to load any file
    try:
        employee_count = db_manager.get_employee_count()
        if employee_count > 0:
            print(f"[auto_load_employee_file] Employees already loaded in database ({employee_count} employees), skipping auto-load")
            return
    except:
        pass  # If get_employee_count fails, continue with auto-load attempt
    
    # Use glob patterns to find employee files (supports flexible naming)
    # Pattern priority: files with "_Emails" are preferred over those without
    glob_patterns = [
        "Employee Headcount*Emails.csv",  # Preferred pattern (e.g., "Employee Headcount 2025_Emails.csv")
        "Employee Headcount*.csv"          # Fallback pattern (e.g., "Employee Headcount 2025.csv")
    ]
    
    employee_file_candidates = []
    seen_files = set()
    for pattern in glob_patterns:
        pattern_path = os.path.join(script_dir, pattern)
        matched_files = glob.glob(pattern_path)
        if matched_files:
            # Sort files in reverse alphabetical order for consistent prioritization
            # Note: This typically puts files with later months/years first (e.g., "Nov" before "Oct")
            # but doesn't guarantee chronological ordering for all date formats
            matched_files.sort(reverse=True)
            # Only add files we haven't seen yet (to avoid duplicates from overlapping patterns)
            for file_path in matched_files:
                if file_path not in seen_files:
                    employee_file_candidates.append(file_path)
                    seen_files.add(file_path)
            print(f"[auto_load_employee_file] Found {len(matched_files)} file(s) matching pattern '{pattern}'")
    
    if not employee_file_candidates:
        print(f"[auto_load_employee_file] No employee files found matching patterns: {glob_patterns}")
        return
    
    print(f"[auto_load_employee_file] Total {len(employee_file_candidates)} candidate file(s) found")
    
    for file_path in employee_file_candidates:
        filename = os.path.basename(file_path)
        print(f"[auto_load_employee_file] Checking for: {file_path}")
        
        if os.path.exists(file_path):
            print(f"[auto_load_employee_file] Found employee file: {filename}")
            try:
                # Check if this file has already been loaded
                # We'll use a simple marker to avoid reloading the same file repeatedly
                # Place marker in the script directory (same as where we look for the CSV)
                marker_file = os.path.join(script_dir, f".{filename}.loaded")
                
                # Check if CSV is newer than marker file (handles file updates)
                should_reload = False
                if os.path.exists(marker_file):
                    csv_mtime = os.path.getmtime(file_path)
                    marker_mtime = os.path.getmtime(marker_file)
                    
                    if csv_mtime > marker_mtime:
                        print(f"[auto_load_employee_file] CSV file is newer than marker - will reload")
                        should_reload = True
                    else:
                        print(f"[auto_load_employee_file] Marker file found and CSV unchanged, trying next candidate...")
                        continue
                else:
                    print(f"[auto_load_employee_file] No marker file found, will load for first time")
                    should_reload = True
                
                if not should_reload:
                    continue
                
                # Read the employee file
                print(f"[auto_load_employee_file] Reading CSV file: {file_path}")
                emp_df = pd.read_csv(file_path, low_memory=False)
                print(f"[auto_load_employee_file] CSV contains {len(emp_df)} rows")
                
                # Map columns - adjust based on the CSV structure
                # Expected columns: Last Name, First Name, Title, Function (department), Status, Email
                if 'Last Name' in emp_df.columns and 'First Name' in emp_df.columns:
                    print(f"[auto_load_employee_file] CSV has expected column structure")
                    # Create normalized dataframe
                    normalized_emp_df = pd.DataFrame({
                        'first_name': emp_df['First Name'],
                        'last_name': emp_df['Last Name'],
                        'email': emp_df.get('Email', None),
                        'title': emp_df.get('Title', ''),
                        'department': emp_df.get('Function', ''),  # Function column maps to department
                        'status': emp_df.get('Status', '')
                    })
                    
                    # Load into database
                    print(f"[auto_load_employee_file] Loading {len(normalized_emp_df)} employees into database...")
                    success, message, count = db_manager.load_employees(normalized_emp_df)
                    
                    if success:
                        print(f"[auto_load_employee_file] ✅ {message}")
                        # Create marker file to indicate successful load
                        with open(marker_file, 'w') as f:
                            f.write(f"Loaded on {datetime.now().isoformat()}\n")
                            f.write(f"Records: {count}\n")
                        print(f"[auto_load_employee_file] Created marker file: {marker_file}")
                        return  # Successfully loaded, exit
                    else:
                        print(f"[auto_load_employee_file] ❌ Error loading employee file: {message}")
                else:
                    print(f"[auto_load_employee_file] ❌ Employee file {filename} has unexpected column structure")
                    print(f"[auto_load_employee_file] Available columns: {list(emp_df.columns)}")
                    
            except Exception as e:
                print(f"[auto_load_employee_file] ❌ Error auto-loading employee file {filename}: {str(e)}")
                traceback.print_exc()
        else:
            print(f"[auto_load_employee_file] File not found: {file_path}")

db, processor, scanner = init_app()

# Department mapping storage file
DEPT_MAPPING_FILE = "department_mappings.json"

def load_department_mappings():
    """Load department mappings from JSON file."""
    if os.path.exists(DEPT_MAPPING_FILE):
        try:
            with open(DEPT_MAPPING_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_department_mappings(mappings):
    """Save department mappings to JSON file."""
    with open(DEPT_MAPPING_FILE, 'w') as f:
        json.dump(mappings, f, indent=2)

def clear_employee_markers():
    """
    Clear all employee file marker files (.loaded files).
    This allows employee files to be re-imported even if they were previously loaded.
    
    Returns:
        int: Number of marker files deleted
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    deleted_count = 0
    
    try:
        # Look for .*.loaded files in the script directory
        for filename in os.listdir(script_dir):
            if filename.endswith('.loaded') and filename.startswith('.'):
                marker_path = os.path.join(script_dir, filename)
                try:
                    os.remove(marker_path)
                    deleted_count += 1
                    print(f"Deleted marker file: {filename}")
                except Exception as e:
                    print(f"Error deleting marker {filename}: {e}")
    except Exception as e:
        print(f"Error clearing employee markers: {e}")
    
    return deleted_count

def clear_and_reset_all():
    """
    Comprehensive reset function that:
    1. Clears all database data
    2. Removes file tracking JSON
    3. Removes all employee marker files
    
    This provides a complete fresh start for all data processing.
    
    Returns:
        tuple: (success: bool, message: str, details: dict)
    """
    details = {
        'database_cleared': False,
        'tracking_reset': False,
        'markers_cleared': 0
    }
    
    try:
        # 1. Clear database
        db_success = db.delete_all_data()
        details['database_cleared'] = db_success
        
        # 2. Reset file tracking
        scanner.reset_all_tracking()
        details['tracking_reset'] = True
        
        # 3. Clear employee markers
        markers_deleted = clear_employee_markers()
        details['markers_cleared'] = markers_deleted
        
        success = db_success and details['tracking_reset']
        
        if success:
            message = f"✅ Complete reset successful: Database cleared, file tracking reset, {markers_deleted} marker file(s) removed"
        else:
            message = f"⚠️ Partial reset: Database={'✓' if db_success else '✗'}, Tracking={'✓' if details['tracking_reset'] else '✗'}, Markers={markers_deleted}"
        
        return success, message, details
        
    except Exception as e:
        return False, f"❌ Error during reset: {str(e)}", details

def force_reload_employee_file():
    """
    Force reload the employee master file by clearing its marker and re-running auto-load.
    Uses glob patterns to find employee files matching common naming patterns.
    
    Returns:
        tuple: (success: bool, message: str)
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Use glob patterns to find employee files (same as auto_load_employee_file)
    glob_patterns = [
        "Employee Headcount*Emails.csv",
        "Employee Headcount*.csv"
    ]
    
    employee_files = []
    seen_files = set()
    for pattern in glob_patterns:
        pattern_path = os.path.join(script_dir, pattern)
        matched_files = glob.glob(pattern_path)
        if matched_files:
            matched_files.sort(reverse=True)
            # Deduplicate files to avoid processing the same file twice
            for file_path in matched_files:
                if file_path not in seen_files:
                    employee_files.append(file_path)
                    seen_files.add(file_path)
    
    # Clear markers for all found employee files
    markers_cleared = 0
    for file_path in employee_files:
        filename = os.path.basename(file_path)
        marker_file = os.path.join(script_dir, f".{filename}.loaded")
        if os.path.exists(marker_file):
            try:
                os.remove(marker_file)
                markers_cleared += 1
                print(f"Cleared marker for {filename}")
            except Exception as e:
                print(f"Error clearing marker for {filename}: {e}")
    
    # Re-run auto-load
    try:
        auto_load_employee_file(db)
        return True, f"✅ Employee file reload initiated ({markers_cleared} marker(s) cleared)"
    except Exception as e:
        return False, f"❌ Error reloading employee file: {str(e)}"

def apply_department_mappings(data, mappings):
    """Apply department mappings to the dataset."""
    if not mappings or data.empty:
        return data
    
    data = data.copy()
    for email, dept in mappings.items():
        if email in data['email'].values:
            data.loc[data['email'] == email, 'department'] = dept
    
    return data

def apply_employee_departments(data, db_manager=None):
    """
    Apply employee master file departments to all data.
    
    This ensures the employee master file is the authoritative source for 
    department assignments for all employees, not just power users.
    
    Args:
        data: DataFrame with usage data
        db_manager: DatabaseManager instance (optional, will use global db if not provided)
        
    Returns:
        DataFrame with employee departments applied
    """
    if data.empty:
        return data
    
    data = data.copy()
    
    # Use provided db_manager or fall back to global db
    database = db_manager if db_manager is not None else db
    
    # Get all employees for efficient lookup
    try:
        employees_df = database.get_all_employees()
        if employees_df.empty:
            return data
        
        # Create lookup dictionaries for fast matching
        email_to_dept = {}
        name_to_dept = {}
        
        for _, emp in employees_df.iterrows():
            # Email-based lookup (primary)
            if pd.notna(emp.get('email')) and emp.get('email'):
                email_to_dept[emp['email'].lower().strip()] = emp.get('department', 'Unknown')
            
            # Name-based lookup (fallback)
            if pd.notna(emp.get('first_name')) and pd.notna(emp.get('last_name')):
                full_name = f"{emp['first_name']} {emp['last_name']}".lower().strip()
                name_to_dept[full_name] = emp.get('department', 'Unknown')
        
        # Apply employee departments to data
        for idx, row in data.iterrows():
            employee_dept = None
            
            # Try email match first
            if pd.notna(row.get('email')) and row.get('email'):
                email_key = row['email'].lower().strip()
                if email_key in email_to_dept:
                    employee_dept = email_to_dept[email_key]
            
            # Try name match if email didn't work
            if not employee_dept and pd.notna(row.get('user_name')) and row.get('user_name'):
                name_key = row['user_name'].lower().strip()
                if name_key in name_to_dept:
                    employee_dept = name_to_dept[name_key]
            
            # Update department if employee found
            if employee_dept:
                data.at[idx, 'department'] = employee_dept
        
        return data
        
    except AttributeError as e:
        # Handle case where db doesn't have employee methods (cache error)
        print(f"AttributeError in apply_employee_departments: {e}")
        return data
    except Exception as e:
        print(f"Error applying employee departments: {e}")
        traceback.print_exc()
        return data

def is_employee_user(email, user_name):
    """
    Check if a user is an employee by email or name.
    
    Args:
        email: User email address
        user_name: User full name
        
    Returns:
        bool: True if user is an employee, False otherwise
    """
    try:
        # First try by email
        if email:
            employee = db.get_employee_by_email(email)
            if employee:
                return True
        
        # Try by name if email lookup fails
        if user_name:
            name_parts = user_name.strip().split()
            if len(name_parts) >= 2:
                first_name = name_parts[0]
                last_name = ' '.join(name_parts[1:])
                employee = db.get_employee_by_name(first_name, last_name)
                if employee:
                    return True
        
        return False
    except AttributeError:
        # Handle cache error - database object missing methods
        # This happens when code is updated while app is running
        return False

def get_employee_for_user(email, user_name):
    """
    Get employee record for a user by email or name.
    
    Args:
        email: User email address
        user_name: User full name
        
    Returns:
        dict or None: Employee record if found
    """
    try:
        # First try by email
        if email:
            employee = db.get_employee_by_email(email)
            if employee:
                return employee
        
        # Try by name if email lookup fails
        if user_name:
            name_parts = user_name.strip().split()
            if len(name_parts) >= 2:
                first_name = name_parts[0]
                last_name = ' '.join(name_parts[1:])
                employee = db.get_employee_by_name(first_name, last_name)
                if employee:
                    return employee
        
        return None
    except AttributeError:
        # Handle cache error - database object missing methods
        # This happens when code is updated while app is running
        return None

# Enhanced CSS with dark mode support
st.markdown("""
<style>
    /* CSS Variables for Light and Dark Mode */
    :root {
        /* Light mode colors (default) */
        --bg-primary: #ffffff;
        --bg-secondary: #f8f9fa;
        --bg-tertiary: #f1f5f9;
        --text-primary: #1e293b;
        --text-secondary: #475569;
        --text-tertiary: #64748b;
        --border-color: #e9ecef;
        --border-color-light: #e2e8f0;
        --border-dashed: #cbd5e1;
        --shadow-sm: rgba(0,0,0,0.05);
        --shadow-md: rgba(0,0,0,0.1);
        --shadow-lg: rgba(0,0,0,0.15);
        
        /* Card backgrounds */
        --card-bg-start: #f8f9fa;
        --card-bg-end: #ffffff;
        
        /* State colors with good contrast */
        --success-bg: #d1fae5;
        --success-text: #065f46;
        --success-border: #10b981;
        
        --warning-bg: #fef3c7;
        --warning-text: #92400e;
        --warning-border: #f59e0b;
        
        --info-bg: #dbeafe;
        --info-text: #1e40af;
        --info-border: #3b82f6;
        
        --error-bg: #fee2e2;
        --error-text: #991b1b;
        --error-border: #ef4444;
    }
    
    /* Dark mode colors */
    @media (prefers-color-scheme: dark) {
        :root {
            --bg-primary: #1e293b;
            --bg-secondary: #334155;
            --bg-tertiary: #475569;
            --text-primary: #f1f5f9;
            --text-secondary: #cbd5e1;
            --text-tertiary: #94a3b8;
            --border-color: #475569;
            --border-color-light: #64748b;
            --border-dashed: #64748b;
            --shadow-sm: rgba(0,0,0,0.3);
            --shadow-md: rgba(0,0,0,0.4);
            --shadow-lg: rgba(0,0,0,0.5);
            
            /* Card backgrounds for dark mode */
            --card-bg-start: #334155;
            --card-bg-end: #1e293b;
            
            /* State colors optimized for dark mode */
            --success-bg: #065f46;
            --success-text: #a7f3d0;
            --success-border: #34d399;
            
            --warning-bg: #78350f;
            --warning-text: #fde68a;
            --warning-border: #fbbf24;
            
            --info-bg: #1e3a8a;
            --info-text: #bfdbfe;
            --info-border: #60a5fa;
            
            --error-bg: #7f1d1d;
            --error-text: #fecaca;
            --error-border: #f87171;
        }
    }
    
    /* Main header with improved typography */
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 1rem;
        padding: 0.5rem 0;
    }
    
    /* Improved metric cards */
    .metric-card {
        background: linear-gradient(135deg, var(--card-bg-start) 0%, var(--card-bg-end) 100%);
        padding: 1.5rem;
        border-radius: 0.75rem;
        margin: 0.5rem 0;
        box-shadow: 0 2px 4px var(--shadow-sm);
        border: 1px solid var(--border-color);
        transition: all 0.3s ease;
        color: var(--text-primary);
    }
    
    .metric-card:hover {
        box-shadow: 0 4px 8px var(--shadow-md);
        transform: translateY(-2px);
    }
    
    /* Tool badges with better styling */
    .tool-badge {
        display: inline-block;
        padding: 0.35rem 0.85rem;
        border-radius: 1.25rem;
        font-size: 0.85rem;
        font-weight: 600;
        margin: 0.25rem;
        box-shadow: 0 2px 4px var(--shadow-md);
        transition: all 0.2s ease;
    }
    
    .tool-badge:hover {
        transform: scale(1.05);
    }
    
    .tool-chatgpt {
        background: linear-gradient(135deg, #10a37f 0%, #0d8c6d 100%);
        color: white;
    }
    
    .tool-blueflame {
        background: linear-gradient(135deg, #4f46e5 0%, #4338ca 100%);
        color: white;
    }
    
    /* Enhanced power user badge */
    .power-user-badge {
        background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%);
        color: white;
        padding: 0.35rem 0.85rem;
        border-radius: 1.25rem;
        font-size: 0.75rem;
        font-weight: 600;
        box-shadow: 0 2px 4px rgba(0,0,0,0.2);
    }
    
    /* Department mapper with improved styling */
    .dept-mapper-container {
        background: linear-gradient(135deg, var(--card-bg-start) 0%, var(--card-bg-end) 100%);
        border: 2px solid var(--border-color);
        border-radius: 0.75rem;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 4px var(--shadow-sm);
        color: var(--text-primary);
    }
    
    /* Enhanced insight cards */
    .insight-card {
        border-left: 4px solid;
        padding: 1.25rem;
        margin: 1rem 0;
        border-radius: 0.5rem;
        box-shadow: 0 2px 4px var(--shadow-sm);
        transition: all 0.3s ease;
    }
    
    .insight-card:hover {
        box-shadow: 0 4px 8px var(--shadow-md);
    }
    
    .insight-success {
        border-color: var(--success-border);
        background: var(--success-bg);
        color: var(--success-text);
    }
    
    .insight-warning {
        border-color: var(--warning-border);
        background: var(--warning-bg);
        color: var(--warning-text);
    }
    
    .insight-info {
        border-color: var(--info-border);
        background: var(--info-bg);
        color: var(--info-text);
    }
    
    /* Enhanced file upload zone */
    .upload-zone {
        border: 2px dashed var(--border-dashed);
        border-radius: 0.75rem;
        padding: 2rem;
        text-align: center;
        background: var(--bg-secondary);
        margin: 1rem 0;
        transition: all 0.3s ease;
        color: var(--text-primary);
    }
    
    .upload-zone:hover {
        border-color: #667eea;
        background: var(--bg-tertiary);
    }
    
    .upload-requirements {
        background: var(--bg-tertiary);
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 0.5rem 0;
        font-size: 0.875rem;
        border: 1px solid var(--border-color-light);
        color: var(--text-primary);
    }
    
    /* Loading states */
    .loading-skeleton {
        background: linear-gradient(90deg, var(--bg-tertiary) 25%, var(--border-color-light) 50%, var(--bg-tertiary) 75%);
        background-size: 200% 100%;
        animation: loading 1.5s ease-in-out infinite;
        border-radius: 0.5rem;
        height: 200px;
    }
    
    @keyframes loading {
        0% { background-position: 200% 0; }
        100% { background-position: -200% 0; }
    }
    
    /* Empty state styling */
    .empty-state {
        text-align: center;
        padding: 3rem 2rem;
        background: var(--bg-secondary);
        border-radius: 0.75rem;
        border: 2px dashed var(--border-dashed);
        margin: 2rem 0;
        color: var(--text-primary);
    }
    
    .empty-state-icon {
        font-size: 4rem;
        margin-bottom: 1rem;
        opacity: 0.5;
    }
    
    .empty-state-title {
        font-size: 1.5rem;
        font-weight: 600;
        color: var(--text-secondary);
        margin-bottom: 0.5rem;
    }
    
    .empty-state-text {
        color: var(--text-tertiary);
        margin-bottom: 1.5rem;
    }
    
    /* Data quality indicators */
    .quality-indicator {
        display: inline-flex;
        align-items: center;
        padding: 0.5rem 1rem;
        border-radius: 0.5rem;
        font-weight: 600;
        font-size: 0.875rem;
        margin: 0.25rem;
    }
    
    .quality-excellent {
        background: var(--success-bg);
        color: var(--success-text);
        border: 1px solid var(--success-border);
    }
    
    .quality-good {
        background: var(--info-bg);
        color: var(--info-text);
        border: 1px solid var(--info-border);
    }
    
    .quality-warning {
        background: var(--warning-bg);
        color: var(--warning-text);
        border: 1px solid var(--warning-border);
    }
    
    .quality-poor {
        background: var(--error-bg);
        color: var(--error-text);
        border: 1px solid var(--error-border);
    }
    
    /* Mobile responsiveness */
    @media (max-width: 768px) {
        .main-header {
            font-size: 1.75rem;
        }
        
        .metric-card {
            padding: 1rem;
        }
        
        .metrics-grid {
            grid-template-columns: 1fr !important;
        }
        
        /* Stack columns on mobile */
        [data-testid="column"] {
            width: 100% !important;
            flex: 100% !important;
            max-width: 100% !important;
        }
    }
    
    @media (max-width: 480px) {
        .main-header {
            font-size: 1.5rem;
        }
        
        body {
            font-size: 14px;
        }
    }
    
    /* Enhanced tooltips */
    .help-tooltip {
        background: var(--info-bg);
        border-left: 4px solid var(--info-border);
        border-radius: 0.5rem;
        padding: 1rem;
        margin: 0.5rem 0;
        font-size: 0.875rem;
        border: 1px solid var(--info-border);
        color: var(--info-text);
    }
    
    .help-tooltip strong {
        color: var(--info-text);
        font-weight: 700;
    }
    
    /* Efficiency badges */
    .efficiency-badge {
        padding: 4px 8px;
        border-radius: 20px;
        font-size: 13px;
        font-weight: 600;
        display: inline-block;
        margin-left: 8px;
    }
    .high {
        background-color: #15803d;
        color: white;
    }
    .medium {
        background-color: #ca8a04;
        color: white;
    }
    .low {
        background-color: #dc2626;
        color: white;
    }
    
    /* Enhanced tab styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: rgba(40, 50, 65, 0.8);
        border-radius: 4px 4px 0px 0px;
        gap: 1px;
        padding: 10px 16px;
        font-weight: 500;
    }
    .stTabs [aria-selected="true"] {
        background-color: rgba(100, 120, 235, 0.2) !important;
        border-bottom: 2px solid #667eea !important;
    }
    
    /* Enhanced dataframe container */
    .dataframe-container {
        border-radius: 10px;
        overflow: hidden;
        background: rgba(35, 45, 60, 0.7);
        padding: 0px;
        margin: 10px 0px;
    }
    
    /* Metric row styling */
    .metric-row {
        display: flex;
        justify-content: space-between;
        border-bottom: 1px solid rgba(100, 116, 139, 0.2);
        padding: 8px 0;
    }
    .metric-label {
        color: #94a3b8;
        font-size: 14px;
    }
    .metric-value {
        font-weight: 600;
        font-size: 15px;
    }
    
    /* Info cards with icons */
    .info-card {
        background: var(--bg-primary);
        border: 2px solid var(--border-color);
        border-radius: 0.75rem;
        padding: 1.25rem;
        margin: 1rem 0;
        box-shadow: 0 1px 3px var(--shadow-md);
        color: var(--text-primary);
    }
    
    .info-card-icon {
        font-size: 2rem;
        margin-bottom: 0.5rem;
    }
    
    /* Responsive tables */
    @media (max-width: 768px) {
        .data-table {
            font-size: 0.875rem;
        }
        
        .data-table th, .data-table td {
            padding: 0.5rem;
        }
    }
    
    /* Enhanced buttons */
    .stButton > button {
        border-radius: 0.5rem;
        font-weight: 600;
        transition: all 0.2s ease;
        box-shadow: 0 2px 4px var(--shadow-md);
    }
    
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 8px var(--shadow-lg);
    }
    
    /* Better spacing */
    .section-header {
        margin-top: 2rem;
        margin-bottom: 1rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid var(--border-color);
        color: var(--text-primary);
    }
</style>
""", unsafe_allow_html=True)

def detect_data_source(df):
    """Detect which AI tool the data is from based on column structure."""
    columns = df.columns.tolist()
    
    # OpenAI ChatGPT detection
    if 'gpt_messages' in columns or 'tool_messages' in columns:
        return 'ChatGPT'
    
    # BlueFlame AI detection - updated for all formats
    # Check for month columns in format 'Mon-YY' (e.g., 'Sep-24', 'Oct-25')
    has_month_cols = any(col for col in columns if len(col.split('-')) == 2 and 
                        col.split('-')[0] in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                             'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec'])
    
    if ('Metric' in columns and any(col.startswith('MoM Var') for col in columns)) or \
       ('Total Messages' in df.values if not df.empty else False) or \
       ('User ID' in columns and has_month_cols) or \
       ('Table' in columns and has_month_cols):
        return 'BlueFlame AI'
    
    # Default or ask user
    return 'Unknown'

def is_weekly_file(filename):
    """
    Detect if a file is a weekly report based on filename.
    Weekly files contain 'weekly' and a date (YYYY-MM-DD format).
    
    Args:
        filename: Name of the file
        
    Returns:
        bool: True if file is detected as weekly report
    """
    filename_lower = filename.lower()
    # Check if filename contains 'weekly' and a date pattern
    import re
    has_weekly = 'weekly' in filename_lower
    has_date = re.search(r'\d{4}-\d{2}-\d{2}', filename) is not None
    return has_weekly and has_date

def determine_record_month(period_start, period_end, first_active, last_active):
    """
    Determine which month a record should be assigned to based on actual usage dates.
    For weekly files that span two months, assign to the month with more activity days.
    
    Args:
        period_start: Period start date (datetime)
        period_end: Period end date (datetime)
        first_active: First day active in period (datetime or None)
        last_active: Last day active in period (datetime or None)
        
    Returns:
        datetime: The date to use for the record (first day of the assigned month)
    """
    # If we have actual activity dates, use them to determine the month
    if pd.notna(first_active) and pd.notna(last_active):
        first_active = pd.to_datetime(first_active, errors='coerce')
        last_active = pd.to_datetime(last_active, errors='coerce')
        
        if pd.notna(first_active) and pd.notna(last_active):
            # Calculate midpoint of actual activity
            midpoint = first_active + (last_active - first_active) / 2
            # Return first day of the month containing the midpoint
            return pd.Timestamp(year=midpoint.year, month=midpoint.month, day=1)
    
    # If no activity dates, use period dates
    if pd.notna(period_start) and pd.notna(period_end):
        period_start = pd.to_datetime(period_start, errors='coerce')
        period_end = pd.to_datetime(period_end, errors='coerce')
        
        if pd.notna(period_start) and pd.notna(period_end):
            # Check if period spans two months
            if period_start.month != period_end.month:
                # Calculate number of days in each month
                days_in_start_month = (pd.Timestamp(year=period_start.year, 
                                                    month=period_start.month, 
                                                    day=1) + pd.DateOffset(months=1) - pd.Timedelta(days=1)).day - period_start.day + 1
                days_in_end_month = period_end.day
                
                # Assign to month with more days
                if days_in_start_month >= days_in_end_month:
                    return pd.Timestamp(year=period_start.year, month=period_start.month, day=1)
                else:
                    return pd.Timestamp(year=period_end.year, month=period_end.month, day=1)
            else:
                # Same month, use period start
                return pd.Timestamp(year=period_start.year, month=period_start.month, day=1)
    
    # Fallback to period_start or current date
    if pd.notna(period_start):
        period_start = pd.to_datetime(period_start, errors='coerce')
        if pd.notna(period_start):
            return pd.Timestamp(year=period_start.year, month=period_start.month, day=1)
    
    # Last resort: current date
    now = datetime.now()
    return pd.Timestamp(year=now.year, month=now.month, day=1)

def normalize_openai_data(df, filename):
    """Normalize OpenAI CSV export to standard schema with enterprise license costs."""
    normalized_records = []
    
    # Get enterprise pricing
    cost_calc = EnterpriseCostCalculator()
    pricing_info = cost_calc.get_pricing_info('ChatGPT')
    monthly_license_cost = pricing_info['license_cost_per_user_monthly']
    
    for _, row in df.iterrows():
        # Get user email and name
        user_email = row.get('email', '')
        user_name = row.get('name', '')
        
        # Look up employee by email to get authoritative department
        employee = None
        try:
            if user_email:
                employee = db.get_employee_by_email(user_email)
            
            # If no match by email, try matching by name
            if not employee and user_name:
                # Try to parse name into first and last
                name_parts = user_name.strip().split()
                if len(name_parts) >= 2:
                    first_name = name_parts[0]
                    last_name = ' '.join(name_parts[1:])  # Handle multi-part last names
                    employee = db.get_employee_by_name(first_name, last_name)
        except AttributeError:
            # Handle cache error - database object missing methods
            employee = None
        
        if employee:
            # Use employee data as source of truth
            dept = employee['department'] if employee['department'] else 'Unknown'
            user_name = f"{employee['first_name']} {employee['last_name']}".strip()
            if not user_name:
                user_name = row.get('name', '')
        else:
            # User not in employee roster - flag as unidentified
            # Parse department - OpenAI exports it as a JSON array string
            dept = 'Unknown'
            user_name = row.get('name', '')
        
        # Get period dates with robust error handling
        period_start = pd.to_datetime(row.get('period_start', row.get('first_day_active_in_period', datetime.now())), errors='coerce')
        period_end = pd.to_datetime(row.get('period_end', row.get('last_day_active_in_period', datetime.now())), errors='coerce')
        first_active = row.get('first_day_active_in_period')
        last_active = row.get('last_day_active_in_period')
        
        # Fallback to current date if parsing fails
        if pd.isna(period_start):
            period_start = datetime.now()
        if pd.isna(period_end):
            period_end = datetime.now()
        
        # Determine the correct month for this record
        # For weekly files spanning two months, this ensures data goes to the right month
        is_weekly = is_weekly_file(filename)
        if is_weekly:
            record_date = determine_record_month(period_start, period_end, first_active, last_active)
        else:
            # For monthly files, use period_start as before
            record_date = period_start
        
        # ChatGPT messages - cost is enterprise license per user per month
        if row.get('messages', 0) > 0:
            normalized_records.append({
                'user_id': row.get('public_id', row.get('email', '')),
                'user_name': user_name,
                'email': user_email,
                'department': dept,
                'date': record_date,
                'feature_used': 'ChatGPT Messages',
                'usage_count': row.get('messages', 0),
                'cost_usd': monthly_license_cost,  # Enterprise license cost per user per month
                'tool_source': 'ChatGPT',
                'file_source': filename
            })
        
        # GPT-specific messages - included in license, cost is 0
        if row.get('gpt_messages', 0) > 0:
            normalized_records.append({
                'user_id': row.get('public_id', row.get('email', '')),
                'user_name': user_name,
                'email': user_email,
                'department': dept,
                'date': record_date,
                'feature_used': 'GPT Messages',
                'usage_count': row.get('gpt_messages', 0),
                'cost_usd': 0,  # Included in base license
                'tool_source': 'ChatGPT',
                'file_source': filename
            })
        
        # Tool messages - included in license, cost is 0
        if row.get('tool_messages', 0) > 0:
            normalized_records.append({
                'user_id': row.get('public_id', row.get('email', '')),
                'user_name': user_name,
                'email': user_email,
                'department': dept,
                'date': record_date,
                'feature_used': 'Tool Messages',
                'usage_count': row.get('tool_messages', 0),
                'cost_usd': 0,  # Included in base license
                'tool_source': 'ChatGPT',
                'file_source': filename
            })
        
        # Project messages - included in license, cost is 0
        if row.get('project_messages', 0) > 0:
            normalized_records.append({
                'user_id': row.get('public_id', row.get('email', '')),
                'user_name': user_name,
                'email': user_email,
                'department': dept,
                'date': record_date,
                'feature_used': 'Project Messages',
                'usage_count': row.get('project_messages', 0),
                'cost_usd': 0,  # Included in base license
                'tool_source': 'ChatGPT',
                'file_source': filename
            })
    
    return pd.DataFrame(normalized_records)

def normalize_blueflame_data(df, filename):
    """Normalize BlueFlame AI data to standard schema with enterprise license costs."""
    normalized_records = []
    
    # Get enterprise pricing for BlueFlame AI
    cost_calc = EnterpriseCostCalculator()
    pricing_info = cost_calc.get_pricing_info('BlueFlame AI')
    monthly_license_cost = pricing_info['license_cost_per_user_monthly']
    
    # Check if this is the combined format with 'Table' column
    if 'Table' in df.columns:
        # Split the dataframe by table type
        monthly_trends = df[df['Table'] == 'Overall Monthly Trends']
        user_data = df[(df['Table'] == 'Top 20 Users Total') | 
                      (df['Table'] == 'Top 10 Increasing Users') | 
                      (df['Table'] == 'Top 10 Decreasing Users')]
        
        # Note: We skip processing monthly trends/aggregate metrics in favor of real user data
        # The user data from Top 20/Top 10 tables provides the actual usage information
        
        # Process user data (from Top 20 Users, Top 10 Increasing, etc.)
        if not user_data.empty:
            # Get month columns (excluding MoM variance columns and non-month columns)
            month_cols = [col for col in user_data.columns if col not in ['Table', 'Rank', 'Metric', 'User ID'] 
                         and not col.startswith('MoM Var')]
            
            # Deduplicate user data - same user may appear in multiple tables (e.g., Top 20 AND Top 10 Increasing)
            # Keep first occurrence for each user
            user_data_deduped = user_data.drop_duplicates(subset=['User ID'], keep='first')
            
            # Process each user
            for _, row in user_data_deduped.iterrows():
                user_email = row['User ID']
                if pd.isna(user_email) or not user_email:
                    continue
                
                # Look up employee by email to get authoritative department
                employee = None
                try:
                    employee = db.get_employee_by_email(user_email) if user_email else None
                    
                    # If no match by email, try to extract name from email and match
                    if not employee and user_email:
                        # Try to parse name from email (e.g., john.doe@company.com -> John Doe)
                        email_name = user_email.split('@')[0].replace('.', ' ').strip()
                        name_parts = email_name.split()
                        if len(name_parts) >= 2:
                            first_name = name_parts[0]
                            last_name = ' '.join(name_parts[1:])
                            employee = db.get_employee_by_name(first_name, last_name)
                except AttributeError:
                    # Handle cache error - database object missing methods
                    employee = None
                
                if employee:
                    # Use employee data as source of truth
                    dept = employee['department'] if employee['department'] else 'Unknown'
                    user_name = f"{employee['first_name']} {employee['last_name']}".strip()
                    if not user_name:
                        user_name = user_email.split('@')[0].replace('.', ' ').title()
                else:
                    # User not in employee roster - flag as unidentified
                    dept = 'Unknown'
                    user_name = user_email.split('@')[0].replace('.', ' ').title()
                
                # Process each month for this user
                for month_col in month_cols:
                    try:
                        # Parse month to a datetime (format like '25-Sep' or 'Sep-25')
                        month_date = None
                        for fmt in ['%y-%b', '%b-%y', '%Y-%b', '%b-%Y']:
                            try:
                                month_date = pd.to_datetime(month_col, format=fmt, errors='coerce')
                                if not pd.isna(month_date):
                                    break
                            except:
                                continue
                        
                        if pd.isna(month_date):
                            continue
                        
                        # Get message count for this month
                        message_count = row[month_col]
                        
                        # Handle dash placeholders and formatting
                        if isinstance(message_count, str):
                            if message_count in ['–', '-', '—', 'N/A', '']:
                                continue
                            message_count = int(message_count.replace(',', ''))
                        
                        # Skip months with no meaningful data
                        if pd.isna(message_count) or message_count == 0:
                            continue
                        
                        # Create user record for this month with enterprise license cost
                        normalized_records.append({
                            'user_id': user_email,
                            'user_name': user_name,
                            'email': user_email,
                            'department': dept,
                            'date': month_date,
                            'feature_used': 'BlueFlame Messages',
                            'usage_count': int(message_count),
                            'cost_usd': monthly_license_cost,  # Enterprise license cost per user per month
                            'tool_source': 'BlueFlame AI',
                            'file_source': filename
                        })
                    except Exception as e:
                        print(f"Error processing month {month_col} for user {user_email}: {str(e)}")
    
    # Check if this is the summary report format with 'Metric' column (but no Table column)
    elif 'Metric' in df.columns and 'User ID' not in df.columns:
        # This format only has aggregate metrics, no individual user data
        # Skip processing as we prefer real user data from other formats
        print("Skipping aggregate-only format without user data")
    
    # If we have the top users file format with User ID column
    elif 'User ID' in df.columns:
        # Get month columns (excluding MoM variance columns)
        month_cols = [col for col in df.columns if col not in ['User ID'] and not col.startswith('MoM Var')]
        
        # Process each user record
        for _, row in df.iterrows():
            user_email = row['User ID']
            user_name = user_email.split('@')[0].replace('.', ' ').title()
            
            # Process each month for this user
            for month_col in month_cols:
                try:
                    # Parse month to a datetime
                    month_date = pd.to_datetime(month_col, format='%b-%y', errors='coerce')
                    if pd.isna(month_date):
                        continue
                    
                    # Get message count for this month
                    message_count = row[month_col]
                    
                    # Handle dash placeholders and formatting
                    if isinstance(message_count, str):
                        if message_count in ['–', '-', '—', 'N/A']:
                            continue
                        message_count = int(message_count.replace(',', ''))
                    
                    # Skip months with no meaningful data
                    if pd.isna(message_count) or message_count == 0:
                        continue
                    
                    # Create user record for this month with enterprise license cost
                    normalized_records.append({
                        'user_id': user_email,
                        'user_name': user_name,
                        'email': user_email,
                        'department': 'BlueFlame Users',  # Default department, can be updated later
                        'date': month_date,
                        'feature_used': 'BlueFlame Messages',
                        'usage_count': message_count,
                        'cost_usd': monthly_license_cost,  # Enterprise license cost per user per month
                        'tool_source': 'BlueFlame AI',
                        'file_source': filename
                    })
                
                except Exception as e:
                    print(f"Error processing month {month_col} for user {user_email}: {str(e)}")
    
    # Other formats (possibly old BlueFlame format or future formats)
    else:
        # Process each user record using best effort detection
        for idx, row in df.iterrows():
            user = row.get('User', row.get('Email', f'blueflame-user-{idx}'))
            email = row.get('Email', f'{user.lower().replace(" ", ".")}@company.com')
            messages = row.get('Messages', row.get('Usage', 0))
            date_col = next((col for col in df.columns if 'Date' in col or 'Month' in col), None)
            
            if date_col:
                try:
                    date = pd.to_datetime(row[date_col], errors='coerce')
                    if pd.isna(date):
                        date = datetime.now()
                except:
                    date = datetime.now()
            else:
                date = datetime.now()
            
            if messages > 0:
                normalized_records.append({
                    'user_id': email,
                    'user_name': user,
                    'email': email,
                    'department': row.get('Department', 'BlueFlame Users'),
                    'date': date,
                    'feature_used': 'BlueFlame Messages',
                    'usage_count': messages,
                    'cost_usd': monthly_license_cost,  # Enterprise license cost per user per month
                    'tool_source': 'BlueFlame AI',
                    'file_source': filename
                })
    
    return pd.DataFrame(normalized_records)

def display_department_mapper():
    """Display department mapping interface with improved user deduplication and pagination."""
    st.subheader("🏢 Department Mapping Tool")
    
    st.markdown("""
    <div class="dept-mapper-container">
    <p><strong>Why use this?</strong> AI tool exports sometimes have incorrect or missing department data. 
    Use this tool to correct department assignments. Changes are saved and applied to all analytics.</p>
    <p><strong>Note:</strong> Employee departments are sourced from the master employee file and are read-only. 
    You can only map departments for unidentified users.</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Load existing mappings
    mappings = load_department_mappings()
    
    # Get all unique users from database
    all_data = db.get_all_data()
    if all_data.empty:
        st.info("No data available. Upload data first to use department mapping.")
        return
    
    # Get unidentified users
    unidentified_users_df = db.get_unidentified_users()
    
    # Show unidentified users section prominently
    if not unidentified_users_df.empty:
        st.markdown('<div class="section-header"><h3>⚠️ Unidentified Users</h3></div>', unsafe_allow_html=True)
        st.warning(f"Found {len(unidentified_users_df)} users not in the employee master file")
        
        with st.expander(f"👥 View {len(unidentified_users_df)} Unidentified Users", expanded=True):
            # Display unidentified users
            for idx, row in unidentified_users_df.iterrows():
                col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
                
                with col1:
                    # Handle NULL/empty user_name
                    display_name = row['user_name'] if pd.notna(row['user_name']) and row['user_name'] else 'Unknown User'
                    st.write(f"**{display_name}**")
                    # Handle NULL/empty email
                    display_email = row['email'] if pd.notna(row['email']) and row['email'] else 'No email'
                    st.caption(display_email)
                
                with col2:
                    # Handle NULL tools_used
                    display_tools = row['tools_used'] if pd.notna(row['tools_used']) and row['tools_used'] else 'Unknown'
                    st.write(f"🔧 {display_tools}")
                
                with col3:
                    # Handle NaN/NULL total_usage
                    if pd.notna(row['total_usage']):
                        st.write(f"📊 {int(row['total_usage']):,} messages")
                    else:
                        st.write("📊 0 messages")
                
                with col4:
                    # Handle NaN/NULL days_active
                    if pd.notna(row['days_active']):
                        st.caption(f"{int(row['days_active'])} days active")
                    else:
                        st.caption("0 days")
                
                st.divider()
        
        st.divider()
    
    # Department options - hardcoded list of 59 departments
    dept_options = [
        'Administrative - Capital Formation',
        'Administrative - Communications',
        'Administrative - Corporate Credit',
        'Administrative - Events & Premises',
        'Administrative - Finance',
        'Administrative - GPS',
        'Administrative - Investments',
        'Administrative - Legal',
        'Administrative - President',
        'Administrative - TBO',
        'Administrative - TMO',
        'Adminstrative - Investments',
        'CEO Office',
        'Capital Formation',
        'Communications',
        'Compliance',
        'Corporate Credit - Asset Based',
        'Corporate Credit - Asset Management',
        'Corporate Credit - Capital Markets',
        'Corporate Credit - Consumer',
        'Corporate Credit - Credit',
        'Corporate Credit - Diversified Credit',
        'Corporate Credit - Healthcare',
        'Corporate Credit - Industrials',
        'Corporate Credit - Manufacturing',
        'Corporate Credit - Media',
        'Corporate Credit - Middle Market',
        'Corporate Credit - Origination',
        'Corporate Credit - Risk & Restructuring',
        'Corporate Credit - Technology',
        'Corporate Credit - Trading',
        'Corporate Credit -Industrials',
        'Cross Asset Specialists',
        'Cross Platform Solutions',
        'Events',
        'Executive',
        'Finance',
        'Finance - Capital Markets',
        'Finance - Controllership',
        'Finance - Diversified Credit',
        'Finance - FP&A',
        'Finance - Tax',
        'GP Solutions',
        'Human Capital ',
        'Investments',
        'Legal',
        'Operations',
        'Operations - Project Management',
        'Ops - Client Service & Invest. Ops ',
        'Ops - Portfolio & Fund Ops',
        'Premises',
        'Private Wealth',
        'Product Development',
        'Real Estate Credit',
        'Real Estate Credit - Portfolio Management',
        'SME',
        'Structured Credit',
        'Technology',
        'Unknown'
    ]
    
    # Deduplicate users by email only, using smart department selection
    # This prevents users who appear in both OpenAI and BlueFlame from showing as duplicates
    users_df = all_data.groupby('email').agg({
        'user_name': 'first',
        'department': lambda x: _select_primary_department(x),
        'tool_source': lambda x: ', '.join(sorted(x.unique()))
    }).reset_index()
    users_df = users_df.sort_values('user_name')
    
    # Check which users are employees (by email or name)
    users_df['is_employee'] = users_df.apply(lambda row: is_employee_user(row['email'], row['user_name']), axis=1)
    
    st.markdown('<div class="section-header"><h3>📋 All Users</h3></div>', unsafe_allow_html=True)
    st.write(f"**Total Users:** {len(users_df)} ({users_df['is_employee'].sum()} employees, {(~users_df['is_employee']).sum()} unidentified)")
    
    # Filter options
    col1, col2 = st.columns(2)
    with col1:
        filter_type = st.radio("Filter by:", ["All Users", "Employees Only", "Unidentified Only"], horizontal=True)
    
    # Apply filter
    if filter_type == "Employees Only":
        users_df = users_df[users_df['is_employee']]
    elif filter_type == "Unidentified Only":
        users_df = users_df[~users_df['is_employee']]
    
    # Search filter
    with col2:
        search = st.text_input("🔍 Search by name or email", "")
    
    # Initialize pagination state if not exists
    if 'dept_mapper_page' not in st.session_state:
        st.session_state.dept_mapper_page = 0
    
    # Users per page
    users_per_page = 20
    
    # Apply search filter if provided
    if search:
        users_df = users_df[
            users_df['user_name'].str.contains(search, case=False, na=False) | 
            users_df['email'].str.contains(search, case=False, na=False)
        ]
        # Reset pagination when searching
        st.session_state.dept_mapper_page = 0
    
    # Calculate total pages
    total_pages = max(1, (len(users_df) + users_per_page - 1) // users_per_page)
    
    # Ensure current page is valid
    st.session_state.dept_mapper_page = min(st.session_state.dept_mapper_page, total_pages - 1)
    st.session_state.dept_mapper_page = max(0, st.session_state.dept_mapper_page)
    
    # Pagination controls - place them at top and bottom
    def pagination_controls(position="top"):
        col1, col2, col3, col4, col5 = st.columns([2, 1, 3, 1, 2])
        
        with col1:
            if st.button("◀️ Previous", key=f"prev_{position}", 
                         disabled=(st.session_state.dept_mapper_page <= 0)):
                st.session_state.dept_mapper_page -= 1
        
        with col3:
            # Show page selector with page numbers
            page_options = [f"Page {i+1} of {total_pages}" for i in range(total_pages)]
            selected_page = st.selectbox(
                "Page",
                options=page_options,
                index=st.session_state.dept_mapper_page,
                key=f"page_select_{position}",
                label_visibility="collapsed"
            )
            if page_options.index(selected_page) != st.session_state.dept_mapper_page:
                st.session_state.dept_mapper_page = page_options.index(selected_page)
        
        with col5:
            if st.button("Next ▶️", key=f"next_{position}", 
                         disabled=(st.session_state.dept_mapper_page >= total_pages - 1)):
                st.session_state.dept_mapper_page += 1
    
    # Display pagination at top
    pagination_controls("top")
    
    # Get current page of users
    start_idx = st.session_state.dept_mapper_page * users_per_page
    end_idx = start_idx + users_per_page
    current_page_users = users_df.iloc[start_idx:end_idx]
    
    # Display in a clean table format with editable departments
    st.write("**Update Department Assignments:**")
    
    changes_made = False
    
    # Create columns for the table
    col1, col2, col3, col4, col5 = st.columns([2, 3, 2, 2, 1])
    with col1:
        st.write("**Name**")
    with col2:
        st.write("**Email**")
    with col3:
        st.write("**Type**")
    with col4:
        st.write("**Department**")
    with col5:
        st.write("**Action**")
    
    st.divider()
    
    # Show current page of users
    for position, (idx, row) in enumerate(current_page_users.iterrows()):
        col1, col2, col3, col4, col5 = st.columns([2, 3, 2, 2, 1])
        
        with col1:
            # Show user name with multi-tool indicator
            # Handle NULL/empty user_name
            user_display = row['user_name'] if pd.notna(row['user_name']) and row['user_name'] else 'Unknown User'
            # Handle NULL/empty tool_source
            tool_source = row['tool_source'] if pd.notna(row['tool_source']) and row['tool_source'] else ''
            if ', ' in tool_source:  # User has multiple AI tools
                user_display = f"🔗 {user_display}"
            st.write(user_display)
        
        with col2:
            # Show email with tool sources on hover
            # Handle NULL/empty email and tool_source
            display_email = row['email'] if pd.notna(row['email']) and row['email'] else 'No email'
            display_tool_help = tool_source if tool_source else 'Unknown'
            st.text(display_email, help=f"Tools: {display_tool_help}")
        
        with col3:
            # Show if employee or unidentified
            if row['is_employee']:
                st.write("✅ Employee")
            else:
                st.write("⚠️ Unidentified")
        
        with col4:
            current_dept = mappings.get(row['email'], row['department'])
            
            # If employee, show department as read-only
            if row['is_employee']:
                employee = get_employee_for_user(row['email'], row['user_name'])
                if employee and employee.get('department'):
                    st.write(f"🔒 {employee['department']}")
                else:
                    # Handle NULL/empty department
                    display_dept = current_dept if pd.notna(current_dept) and current_dept else 'Unknown'
                    st.write(display_dept)
            else:
                # For unidentified users, allow editing
                new_dept = st.selectbox(
                    "Dept",
                    options=dept_options,
                    index=dept_options.index(current_dept) if current_dept in dept_options else dept_options.index('Unknown'),
                    key=f"dept_{start_idx + position}_{row['email']}",
                    label_visibility="collapsed"
                )
                
                if new_dept != mappings.get(row['email'], row['department']):
                    mappings[row['email']] = new_dept
                    changes_made = True
        
        with col5:
            if not row['is_employee']:
                if st.button("✓", key=f"save_{start_idx + position}_{row['email']}", help="Save changes"):
                    changes_made = True
        
        st.divider()
    
    # Display pagination at bottom too
    pagination_controls("bottom")
    
    # User info message
    showing_text = f"Showing users {start_idx + 1}-{min(end_idx, len(users_df))} of {len(users_df)}"
    if search:
        showing_text += f" (filtered by '{search}')"
    st.info(showing_text)
    
    # Save button
    if changes_made or st.button("💾 Save All Department Mappings", type="primary"):
        save_department_mappings(mappings)
        st.success(f"✅ Saved department mappings for {len(mappings)} users")
    
    # Show current mappings count
    if mappings:
        st.info(f"📊 {len(mappings)} custom department mappings active")

def process_auto_file(file_info, tool_type='Auto-Detect'):
    """
    Process a file from auto-scan folders.
    
    Args:
        file_info: Dictionary with file information from FileScanner
        tool_type: Type of AI tool (Auto-Detect, OpenAI ChatGPT, etc.)
        
    Returns:
        tuple: (success: bool, message: str, records_count: int)
    """
    try:
        file_path = file_info['path']
        
        # Read file from filesystem
        df, read_error = read_file_from_path(file_path)
        
        if read_error:
            return False, f"Error reading file: {read_error}", 0
        
        if df is None or df.empty:
            return False, "File contains no data", 0
        
        # Detect data source
        if tool_type == 'Auto-Detect':
            detected_tool = detect_data_source(df)
        else:
            detected_tool = tool_type.replace('OpenAI ', '')
        
        # Normalize data based on detected tool
        if 'ChatGPT' in detected_tool:
            normalized_df = normalize_openai_data(df, file_info['filename'])
        elif 'BlueFlame' in detected_tool:
            normalized_df = normalize_blueflame_data(df, file_info['filename'])
        else:
            return False, f"Unknown data format: {detected_tool}", 0
        
        # Process the normalized data
        if not normalized_df.empty:
            success, message = processor.process_monthly_data(normalized_df, file_info['filename'])
            
            if success:
                # Mark file as processed
                scanner.mark_processed(
                    file_path, 
                    success=True, 
                    records_count=len(normalized_df)
                )
                return True, f"Successfully processed {len(normalized_df)} records", len(normalized_df)
            else:
                scanner.mark_processed(
                    file_path, 
                    success=False, 
                    error=message
                )
                return False, message, 0
        else:
            return False, "No valid data found after normalization", 0
    
    except Exception as e:
        error_msg = f"Error processing file: {str(e)}"
        scanner.mark_processed(
            file_info['path'], 
            success=False, 
            error=error_msg
        )
        return False, error_msg, 0

def calculate_power_users(data, threshold_percentile=95):
    """Identify power users based on usage patterns."""
    if data.empty:
        return pd.DataFrame()
    
    # Group by email only to avoid duplicates across different departments/tools
    user_usage = data.groupby('email').agg({
        'user_name': 'first',  # Take first user_name (should be same for same email)
        'usage_count': 'sum',
        'cost_usd': 'sum',
        'tool_source': lambda x: ', '.join(sorted(x.unique())),
        'department': lambda x: _select_primary_department(x)
    }).reset_index()
    
    # Update departments from employee database as authoritative source
    # This ensures verified employees show their correct locked departments
    for idx, row in user_usage.iterrows():
        employee = get_employee_for_user(row['email'], row['user_name'])
        if employee and employee.get('department'):
            user_usage.at[idx, 'department'] = employee['department']
    
    # Calculate threshold (top 5% by default)
    threshold = user_usage['usage_count'].quantile(threshold_percentile / 100)
    
    # Identify power users (top 5% by usage)
    power_users = user_usage[
        user_usage['usage_count'] >= threshold
    ].sort_values('usage_count', ascending=False)
    
    return power_users

def _select_primary_department(departments):
    """Select the most appropriate department from a list.
    
    Prefers valid department names over placeholder values like 'BlueFlame Users' 
    or 'Unknown'. If multiple valid departments exist, returns the first one.
    Returns placeholder values only if that's all that's available.
    """
    unique_depts = list(departments.unique())
    
    # Filter out placeholder departments if real departments exist
    real_depts = [d for d in unique_depts if d not in ('BlueFlame Users', 'Unknown')]
    
    if real_depts:
        return real_depts[0]  # Return first real department
    
    # If only placeholders available, prefer 'BlueFlame Users' over 'Unknown'
    if 'BlueFlame Users' in unique_depts:
        return 'BlueFlame Users'
    
    return unique_depts[0] if unique_depts else 'Unknown'

def get_user_message_breakdown(data, email):
    """Get message type breakdown for a specific user, organized by tool source."""
    user_data = data[data['email'] == email]
    
    breakdown = {
        'openai': {
            'ChatGPT Messages': 0,
            'GPT Messages': 0,
            'Tool Messages': 0,
            'Project Messages': 0
        },
        'blueflame': {
            'BlueFlame Messages': 0
        }
    }
    
    if not user_data.empty:
        # Get counts grouped by feature type
        message_counts = user_data.groupby('feature_used')['usage_count'].sum().to_dict()
        
        # Map to breakdown structure
        for msg_type, count in message_counts.items():
            if msg_type in breakdown['openai']:
                breakdown['openai'][msg_type] = count
            elif msg_type == 'BlueFlame Messages':
                breakdown['blueflame']['BlueFlame Messages'] = count
    
    return breakdown

def get_department_message_breakdown(data, department):
    """Get message type breakdown for a specific department."""
    dept_data = data[data['department'] == department]
    
    breakdown = {}
    
    if not dept_data.empty:
        # Get counts grouped by feature type
        message_counts = dept_data.groupby('feature_used')['usage_count'].sum().to_dict()
        breakdown = message_counts
    
    return breakdown

def get_all_message_types(data):
    """Get all unique message types from the data."""
    if data.empty or 'feature_used' not in data.columns:
        return []
    return sorted(data['feature_used'].unique().tolist())

def get_organization_message_breakdown(data):
    """Get message type breakdown for the entire organization."""
    breakdown = {}
    
    if not data.empty and 'feature_used' in data.columns:
        # Get counts grouped by feature type
        message_counts = data.groupby('feature_used')['usage_count'].sum().to_dict()
        breakdown = message_counts
    
    return breakdown

def format_message_breakdown_text(breakdown_dict):
    """Format a message breakdown dictionary into readable text."""
    if not breakdown_dict:
        return "No data"
    
    parts = []
    for msg_type, count in sorted(breakdown_dict.items()):
        if count > 0:
            parts.append(f"{msg_type}: {count:,}")
    
    return " | ".join(parts) if parts else "No messages"

def display_tool_comparison(data):
    """Display side-by-side tool comparison."""
    st.subheader("🔄 Tool Comparison View")
    
    # Get tool breakdown
    tools = data['tool_source'].unique()
    
    if len(tools) < 2:
        st.info("Upload data from multiple AI tools to see comparison metrics.")
        return
    
    # Create comparison columns
    cols = st.columns(len(tools))
    
    for idx, tool in enumerate(tools):
        tool_data = data[data['tool_source'] == tool]
        
        with cols[idx]:
            # Tool header with badge
            badge_class = f"tool-{tool.lower().replace(' ', '')}"
            st.markdown(f'<div class="tool-badge {badge_class}">{tool}</div>', unsafe_allow_html=True)
            
            # Key metrics - USAGE ONLY, NO COSTS
            active_users = tool_data['user_id'].nunique()
            total_usage = tool_data['usage_count'].sum()
            
            st.metric("Active Users", f"{active_users}")
            st.metric("Total Messages", f"{total_usage:,}")
            st.metric("Messages per User", f"{total_usage / max(active_users, 1):.0f}")
    
    st.divider()
    
    # Overlap analysis
    st.subheader("🔗 User Overlap Analysis")
    
    user_tools = data.groupby('email')['tool_source'].apply(set).reset_index()
    multi_tool_users = user_tools[user_tools['tool_source'].apply(len) > 1]
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total Unique Users", data['email'].nunique())
    
    with col2:
        st.metric("Multi-Tool Users", len(multi_tool_users))
    
    with col3:
        overlap_pct = (len(multi_tool_users) / max(data['email'].nunique(), 1)) * 100
        st.metric("Overlap Rate", f"{overlap_pct:.1f}%")
    
    # Insights
    st.subheader("💡 Strategic Insights")
    
    if overlap_pct > 50:
        st.markdown("""
        <div class="insight-card insight-warning">
        <strong>⚠️ High Tool Overlap Detected</strong><br>
        Over 50% of users are using multiple AI tools. Consider:
        <ul>
            <li>Surveying users about distinct use cases for each tool</li>
            <li>Evaluating if tools have redundant capabilities</li>
            <li>Understanding workflow patterns across tools</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="insight-card insight-success">
        <strong>✅ Healthy Tool Distribution</strong><br>
        Tools appear to serve distinct user needs with minimal overlap.
        </div>
        """, unsafe_allow_html=True)
    
    # Usage comparison by volume (NOT cost)
    tool_usage = data.groupby('tool_source')['usage_count'].sum().reset_index()
    
    fig = px.bar(
        tool_usage,
        x='tool_source',
        y='usage_count',
        title='Message Volume by Tool',
        labels={'tool_source': 'Tool', 'usage_count': 'Total Messages'},
        color='tool_source'
    )
    st.plotly_chart(fig, use_container_width=True)

def get_openai_data(data):
    """
    Filter dataset to only OpenAI/ChatGPT records.
    
    Args:
        data: Full dataset DataFrame
        
    Returns:
        DataFrame with only OpenAI records
    """
    if data.empty or 'tool_source' not in data.columns:
        return pd.DataFrame()
    
    # Filter to ChatGPT/OpenAI only
    openai_data = data[data['tool_source'].isin(['ChatGPT', 'OpenAI'])].copy()
    
    return openai_data


def calculate_duau(data):
    """
    Calculate Daily Unique Active Users for OpenAI data.
    
    Args:
        data: OpenAI-filtered DataFrame
        
    Returns:
        float: Average daily unique active users
    """
    if data.empty or 'date' not in data.columns:
        return 0
    
    # Group by date and count unique users
    daily_users = data.groupby('date')['user_id'].nunique()
    
    # Calculate average
    duau = daily_users.mean() if not daily_users.empty else 0
    
    return duau


def calculate_days_active_per_month(data):
    """
    Calculate average days active per month for each user.
    
    Args:
        data: OpenAI-filtered DataFrame
        
    Returns:
        DataFrame with user-level days active statistics
    """
    if data.empty or 'date' not in data.columns:
        return pd.DataFrame()
    
    # Parse dates
    data_copy = data.copy()
    data_copy['date'] = pd.to_datetime(data_copy['date'], errors='coerce')
    data_copy = data_copy.dropna(subset=['date'])
    
    # Extract year-month
    data_copy['year_month'] = data_copy['date'].dt.to_period('M')
    
    # Count unique days per user per month
    user_days = data_copy.groupby(['user_id', 'user_name', 'year_month']).agg({
        'date': 'nunique'
    }).reset_index()
    user_days.columns = ['user_id', 'user_name', 'year_month', 'days_active']
    
    # Calculate average days active per month per user
    avg_days = user_days.groupby(['user_id', 'user_name'])['days_active'].mean().reset_index()
    avg_days.columns = ['user_id', 'user_name', 'avg_days_per_month']
    
    return avg_days


def get_user_activity_tiers(data):
    """
    Categorize users into activity tiers based on message volume.
    
    Tiers:
    - Heavy: Top 20% of users by message count
    - Moderate: Next 30% of users
    - Light: Next 30% of users  
    - Inactive: Bottom 20% of users
    
    Args:
        data: OpenAI-filtered DataFrame
        
    Returns:
        DataFrame with user tiers including activity dates
    """
    if data.empty:
        return pd.DataFrame()
    
    # Get total messages per user along with activity dates and department
    agg_dict = {
        'usage_count': 'sum'
    }
    
    # Add activity date fields if they exist
    if 'last_day_active' in data.columns:
        agg_dict['last_day_active'] = 'max'
    if 'first_day_active_in_period' in data.columns:
        agg_dict['first_day_active_in_period'] = 'min'
    if 'last_day_active_in_period' in data.columns:
        agg_dict['last_day_active_in_period'] = 'max'
    if 'department' in data.columns:
        agg_dict['department'] = 'first'
    
    group_cols = ['user_id', 'user_name', 'email']
    user_totals = data.groupby(group_cols).agg(agg_dict).reset_index()
    
    # Rename usage_count to total_messages
    user_totals.rename(columns={'usage_count': 'total_messages'}, inplace=True)
    
    # Get breakdown by feature type
    feature_breakdown = data.groupby(['user_id', 'feature_used'])['usage_count'].sum().reset_index()
    
    # Pivot to get messages, tool_messages, project_messages columns
    feature_pivot = feature_breakdown.pivot(index='user_id', columns='feature_used', values='usage_count').reset_index()
    feature_pivot.columns.name = None
    
    # Merge feature breakdown with user totals
    user_totals = user_totals.merge(feature_pivot, on='user_id', how='left')
    
    # Fill NaN values with 0 for message counts
    for col in ['ChatGPT Messages', 'GPT Messages', 'Tool Messages', 'Project Messages']:
        if col in user_totals.columns:
            user_totals[col] = user_totals[col].fillna(0).astype(int)
        else:
            user_totals[col] = 0
    
    # Calculate percentiles
    p80 = user_totals['total_messages'].quantile(0.80)
    p50 = user_totals['total_messages'].quantile(0.50)
    p20 = user_totals['total_messages'].quantile(0.20)
    
    # Assign tiers
    def assign_tier(messages):
        if messages >= p80:
            return 'Heavy'
        elif messages >= p50:
            return 'Moderate'
        elif messages >= p20:
            return 'Light'
        else:
            return 'Inactive'
    
    user_totals['activity_tier'] = user_totals['total_messages'].apply(assign_tier)
    
    return user_totals


def get_feature_adoption_timeline(data):
    """
    Get timeline of feature adoption (when users started using each feature).
    
    Args:
        data: OpenAI-filtered DataFrame
        
    Returns:
        DataFrame with first usage date per feature per user
    """
    if data.empty or 'feature_used' not in data.columns:
        return pd.DataFrame()
    
    # Parse dates
    data_copy = data.copy()
    data_copy['date'] = pd.to_datetime(data_copy['date'], errors='coerce')
    data_copy = data_copy.dropna(subset=['date'])
    
    # Get first usage date per feature per user
    first_usage = data_copy.groupby(['user_id', 'user_name', 'feature_used']).agg({
        'date': 'min'
    }).reset_index()
    first_usage.columns = ['user_id', 'user_name', 'feature', 'first_used_date']
    
    return first_usage


def calculate_weekly_trends(data):
    """
    Calculate weekly trends for OpenAI data.
    
    Args:
        data: OpenAI-filtered DataFrame
        
    Returns:
        DataFrame with weekly aggregated metrics
    """
    if data.empty or 'date' not in data.columns:
        return pd.DataFrame()
    
    # Parse dates
    data_copy = data.copy()
    data_copy['date'] = pd.to_datetime(data_copy['date'], errors='coerce')
    data_copy = data_copy.dropna(subset=['date'])
    
    # Extract week - keep period for grouping
    data_copy['week'] = data_copy['date'].dt.to_period('W')
    
    # Aggregate by week
    weekly = data_copy.groupby('week').agg({
        'user_id': 'nunique',
        'usage_count': 'sum'
    }).reset_index()
    weekly.columns = ['week', 'active_users', 'total_messages']
    
    # Format week for display as MM/DD/YYYY (vectorized)
    weekly['week_display'] = weekly['week'].dt.start_time.dt.strftime(WEEKLY_CHART_DATE_FORMAT)
    
    return weekly

def main():
    # Main header - professional title without emoji
    col1, col2 = st.columns([4, 1])
    with col1:
        st.markdown('<h1 class="main-header">AI Usage Analytics Dashboard</h1>', unsafe_allow_html=True)
        st.caption("Multi-Platform Analytics for Enterprise AI Tools")
    with col2:
        # Compact actions menu in top-right
        st.markdown('<div style="text-align: right; padding-top: 0.5rem;"></div>', unsafe_allow_html=True)
    
    # Create main tabs
    # NOTE: Keep tab count and variable count in sync to avoid unpacking errors
    tab1, tab2, tab_openai, tab3, tab4, tab5, tab_roi, tab6 = st.tabs([
        "📊 Executive Overview", 
        "🔄 Tool Comparison",
        "🤖 OpenAI Analytics",
        "⭐ Power Users",
        "📈 Message Type Analytics",
        "💎 ROI Analytics",
        "🏢 Department Mapper",
        "🔧 Database Management"
    ])
    
    # Sidebar
    with st.sidebar:
        st.header("🔧 Controls")
        
        # Enhanced file upload section
        st.subheader("📁 Upload Data")
        
        # File requirements info box
        st.markdown("""
        <div class="upload-requirements">
            <strong>📋 Accepted Formats:</strong><br>
            • CSV files (.csv)<br>
            • Excel files (.xlsx)<br>
            <strong>📏 Max Size:</strong> 200MB
        </div>
        """, unsafe_allow_html=True)
        
        # Tool selector with enhanced help
        tool_type = st.selectbox(
            "Select AI Tool",
            options=['Auto-Detect', 'OpenAI ChatGPT', 'BlueFlame AI', 'Other'],
            help="💡 Auto-Detect will analyze your file and identify the data source automatically"
        )
        
        uploaded_file = st.file_uploader(
            "Upload Usage Data (CSV/Excel)",
            type=['csv', 'xlsx'],
            help="🔼 Drag and drop your file here or click to browse"
        )
        
        if uploaded_file is not None:
            # Show file info with enhanced styling
            file_size_mb = uploaded_file.size / (1024 * 1024)
            st.success(f"✅ File loaded: **{uploaded_file.name}** ({file_size_mb:.2f} MB)")
            
            # Show file preview with better error handling
            try:
                with st.spinner("🔍 Reading file preview..."):
                    preview_df, preview_error = read_file_robust(uploaded_file, nrows=5)
                    
                    if preview_error:
                        display_file_error(preview_error)
                    else:
                        st.write("**📊 File Preview:**")
                        st.dataframe(preview_df.head(3), height=120)
                        
                        # Enhanced file statistics
                        col1, col2 = st.columns(2)
                        with col1:
                            st.metric("Columns", len(preview_df.columns))
                        with col2:
                            # Get full row count
                            try:
                                full_df, full_error = read_file_robust(uploaded_file)
                                if full_error:
                                    st.metric("Rows", "~")
                                else:
                                    st.metric("Rows", len(full_df))
                            except:
                                st.metric("Rows", "~")
                
            except Exception as e:
                st.error(f"❌ Cannot preview file: {str(e)}")
                st.info("💡 Please ensure your file is a valid CSV or Excel file")
            
            if st.button("🚀 Process Upload", type="primary", use_container_width=True):
                # Initialize progress tracking
                progress_bar = st.progress(0)
                status_text = st.empty()
                
                try:
                    # Step 1: Reading file
                    status_text.text("📖 Reading file...")
                    progress_bar.progress(20)
                    
                    df, read_error = read_file_robust(uploaded_file)
                    
                    if read_error:
                        progress_bar.empty()
                        status_text.empty()
                        display_file_error(read_error)
                        return
                    
                    if df is None or df.empty:
                        progress_bar.empty()
                        status_text.empty()
                        st.error("❌ The uploaded file contains no data")
                        return
                    
                    # Step 2: Detecting data source
                    status_text.text("🔍 Detecting data source...")
                    progress_bar.progress(40)
                    
                    if tool_type == 'Auto-Detect':
                        detected_tool = detect_data_source(df)
                        st.info(f"📡 Detected: **{detected_tool}**")
                    else:
                        detected_tool = tool_type.replace('OpenAI ', '')
                    
                    # Step 3: Normalizing data
                    status_text.text("⚙️ Normalizing data structure...")
                    progress_bar.progress(60)
                    
                    if 'ChatGPT' in detected_tool:
                        normalized_df = normalize_openai_data(df, uploaded_file.name)
                    elif 'BlueFlame' in detected_tool:
                        normalized_df = normalize_blueflame_data(df, uploaded_file.name)
                    else:
                        progress_bar.empty()
                        status_text.empty()
                        st.error("❌ Unknown data format. Please select the correct tool.")
                        return
                    
                    # Step 4: Storing in database
                    status_text.text("💾 Storing in database...")
                    progress_bar.progress(80)
                    
                    if not normalized_df.empty:
                        success = processor.process_monthly_data(normalized_df, uploaded_file.name)
                        
                        progress_bar.progress(100)
                        
                        if success:
                            status_text.empty()
                            progress_bar.empty()
                            
                            # Success message with details
                            st.success(f"""
                            ✅ **Upload Complete!**
                            - Processed **{len(normalized_df)}** records
                            - Source: **{detected_tool}**
                            - File: {uploaded_file.name}
                            """)
                            
                            # Show summary metrics
                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Users Found", normalized_df['user_id'].nunique())
                            with col2:
                                st.metric("Total Usage", f"{normalized_df['usage_count'].sum():,}")
                            
                            st.balloons()
                            st.rerun()
                        else:
                            progress_bar.empty()
                            status_text.empty()
                            st.error("❌ Error storing data in database")
                    else:
                        progress_bar.empty()
                        status_text.empty()
                        st.error("❌ No data could be extracted from file")
                        st.info("💡 Please verify your file format matches the selected tool type")
                        
                except Exception as e:
                    progress_bar.empty()
                    status_text.empty()
                    st.error(f"❌ **Error processing file:** {str(e)}")
                    
                    # Provide helpful error context
                    with st.expander("🔧 Troubleshooting Tips"):
                        st.markdown("""
                        **Common issues:**
                        - Ensure your CSV/Excel file is not corrupted
                        - Check that column names match the expected format
                        - Verify the file contains the required data fields
                        - Try selecting the tool type manually instead of Auto-Detect
                        """)
                    st.exception(e)
        
        st.divider()
        
        # Auto-scan section for files in folders
        st.subheader("📂 Auto-Detect Files")
        
        # Scan for files
        detected_files = scanner.scan_folders(AUTO_SCAN_FOLDERS)
        
        if detected_files:
            # Show summary stats
            new_files = [f for f in detected_files if f['status'] in ['new', 'modified']]
            processed_files = [f for f in detected_files if f['status'] == 'processed']
            
            col1, col2 = st.columns(2)
            with col1:
                st.metric("📁 Total Files", len(detected_files))
            with col2:
                st.metric("🆕 New Files", len(new_files))
            
            # Refresh button and Force Reprocess All
            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("🔄 Refresh Files", use_container_width=True):
                    st.rerun()
            
            with col_btn2:
                if st.button("🔄 Force Reprocess All", 
                           help="Reset all file statuses - mark all files as new for reprocessing",
                           use_container_width=True,
                           type="secondary"):
                    if st.session_state.get('confirm_reprocess_all'):
                        # Reset all files in auto-scan folders
                        scanner.reset_all_files_status(AUTO_SCAN_FOLDERS)
                        st.session_state.confirm_reprocess_all = False
                        st.success("✅ All files reset - they will now appear as new")
                        st.rerun()
                    else:
                        st.session_state.confirm_reprocess_all = True
                        st.warning("⚠️ Click again to confirm reset of all file statuses")
            
            # Warning for force reprocess
            if st.session_state.get('confirm_reprocess_all'):
                st.info("💡 This will mark all files as 'new' so they can be reprocessed. No data will be deleted.")
            
            # Show new/unprocessed files
            if new_files:
                with st.expander(f"🆕 New Files ({len(new_files)})", expanded=True):
                    for file_info in new_files:
                        st.markdown(f"""
                        **{file_info['filename']}**  
                        📁 Folder: {file_info['folder']} | 📊 {file_info['size_mb']} MB
                        """)
                        
                        # Process button for individual file
                        if st.button(f"▶️ Process", key=f"process_{file_info['path']}", use_container_width=True):
                            with st.spinner(f"Processing {file_info['filename']}..."):
                                success, message, records = process_auto_file(file_info, tool_type)
                                
                                if success:
                                    st.success(f"✅ {message}")
                                    st.rerun()
                                else:
                                    st.error(f"❌ {message}")
                        
                        st.divider()
                
                # Batch process button for all new files
                if len(new_files) > 1:
                    if st.button(f"⚡ Process All {len(new_files)} New Files", type="primary", use_container_width=True):
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        processed_count = 0
                        total_records = 0
                        errors = []
                        
                        for i, file_info in enumerate(new_files):
                            status_text.text(f"Processing {file_info['filename']}...")
                            progress_bar.progress((i + 1) / len(new_files))
                            
                            success, message, records = process_auto_file(file_info, tool_type)
                            
                            if success:
                                processed_count += 1
                                total_records += records
                            else:
                                errors.append(f"{file_info['filename']}: {message}")
                        
                        progress_bar.empty()
                        status_text.empty()
                        
                        if processed_count > 0:
                            st.success(f"""
                            ✅ **Batch Processing Complete!**
                            - Processed: {processed_count}/{len(new_files)} files
                            - Total Records: {total_records:,}
                            """)
                            
                            if errors:
                                with st.expander(f"⚠️ {len(errors)} Errors"):
                                    for error in errors:
                                        st.error(error)
                            
                            st.balloons()
                            st.rerun()
                        else:
                            st.error("❌ Failed to process any files")
                            for error in errors:
                                st.error(error)
            
            # Show processed files
            if processed_files:
                with st.expander(f"✅ Processed Files ({len(processed_files)})"):
                    for file_info in processed_files:
                        col1, col2 = st.columns([4, 1])
                        
                        with col1:
                            st.markdown(f"""
                            **{file_info['filename']}**  
                            📁 {file_info['folder']} | 📊 {file_info['size_mb']} MB | ✅ {file_info.get('last_processed', 'Unknown')}
                            """)
                        
                        with col2:
                            # Reset button for individual file
                            if st.button("🔄 Reset", key=f"reset_{file_info['path']}", 
                                       help=f"Reset {file_info['filename']} to allow reprocessing",
                                       use_container_width=True):
                                scanner.reset_file_status(file_info['path'])
                                st.success(f"✅ Reset {file_info['filename']}")
                                st.rerun()
                        
                        st.divider()
        else:
            st.info("""
            📂 **No files detected in scan folders**
            
            Place your CSV files in these folders:
            - `OpenAI User Data/`
            - `BlueFlame User Data/`
            - `data/uploads/`
            
            Then click Refresh to detect them.
            """)
        
        # Help Section
        with st.expander("❓ Metrics Guide", expanded=False):
            st.markdown("""
            ### 📊 Understanding Key Metrics
            
            **Total Active Users**
            - Unique users with activity across all AI platforms
            - Indicates reach and adoption of AI tools
            
            **Total Messages**
            - All interactions across AI platforms
            - Key indicator of overall AI engagement
            
            **Messages per User**
            - Average engagement per user
            - Higher values indicate more active tool utilization
            
            **Active Departments**
            - Number of departments using AI tools
            - Measures organizational adoption breadth
            
            **Cost Efficiency**
            - Average cost per message/interaction
            - Lower values indicate better efficiency
            
            **Power Users**
            - Top 5% of users by total usage
            - Key candidates for feedback and beta testing
            
            **Month-over-Month Growth**
            - % change in users, usage, or cost vs previous month
            - Positive values show adoption growth
            
            ### 📁 Data Sources
            
            **OpenAI ChatGPT**
            - Enterprise usage exports
            - Includes: ChatGPT Messages, Tool Messages, Project Messages, GPT Messages
            
            **BlueFlame AI**
            - Custom AI platform usage
            - Monthly message counts per user
            
            ### 🔄 Export Options
            
            **PDF Report (HTML)**
            - Executive summary with key metrics
            - Download and print to PDF from browser
            
            **Excel with Pivots**
            - Multiple sheets with analysis
            - User, Department, Monthly, and Feature summaries
            """)
        
        st.divider()
        
        # Date range and filters
        st.subheader("📅 Filters")
        
        try:
            min_date, max_date = db.get_date_range()
        except:
            # Enhanced empty state for no data
            st.markdown("""
            <div class="empty-state">
                <div class="empty-state-icon">📊</div>
                <div class="empty-state-title">No Data Available</div>
                <div class="empty-state-text">
                    Upload your first AI usage export file to start analyzing data
                </div>
            </div>
            """, unsafe_allow_html=True)
            
            # Helpful guidance
            with st.expander("📚 Getting Started Guide"):
                st.markdown("""
                **To get started:**
                1. Export your AI tool usage data (CSV or Excel)
                2. Select the tool type or use Auto-Detect
                3. Upload your file using the uploader above
                4. Wait for processing to complete
                
                **Supported Tools:**
                - OpenAI ChatGPT Enterprise exports
                - BlueFlame AI usage reports
                - Other AI platforms (manual mapping)
                """)
            return
        
        if min_date and max_date:
            # Enhanced data availability indicator
            date_diff = (max_date - min_date).days
            st.markdown(f"""
            <div class="help-tooltip">
                📊 <strong>Data Available:</strong> {min_date} to {max_date} ({date_diff + 1} days)
            </div>
            """, unsafe_allow_html=True)
            
            date_range = st.date_input(
                "Select date range",
                value=(min_date, max_date),
                min_value=min_date,
                max_value=max_date,
                help="📅 Filter data by date range for focused analysis"
            )
            
            # Enhanced data provider selector
            st.markdown("""
            <div class="help-tooltip">
                🔧 <strong>Data Provider Filter:</strong> Toggle between AI platforms to analyze specific tool usage
            </div>
            """, unsafe_allow_html=True)
            
            all_data = db.get_all_data()
            if not all_data.empty and 'tool_source' in all_data.columns:
                available_tools = list(all_data['tool_source'].unique())
                
                # Create toggle buttons for providers
                st.write("**Select Data Provider:**")
                
                # Use radio buttons for clear selection
                selected_tool = st.radio(
                    "Provider",
                    options=['All Tools'] + available_tools,
                    index=0,
                    help="📊 Filter dashboard to show data from specific AI platform",
                    label_visibility="collapsed"
                )
                
                # Show provider-specific stats
                if selected_tool != 'All Tools':
                    tool_data = all_data[all_data['tool_source'] == selected_tool]
                    st.info(f"📈 {selected_tool}: {tool_data['user_id'].nunique()} users, {len(tool_data):,} records")
            else:
                selected_tool = 'All Tools'
            
            # Department filter with count
            departments = db.get_unique_departments()
            selected_depts = st.multiselect(
                f"Departments ({len(departments)} total)",
                departments,
                help="🏢 Filter by specific departments (leave empty for all)"
            )
            
            # Frequency toggle
            st.divider()
            st.markdown("**📊 Analysis Frequency**")
            freq = st.radio(
                "Frequency",
                ["Monthly (default)", "Weekly"],
                index=0,
                help="Monthly: aggregate OpenAI weeks to months; Weekly: estimate Blueflame weeks from monthly totals."
            )
            
            # Partial period exclusion
            exclude_partial = st.checkbox(
                "Exclude current in-progress period",
                value=True,
                help="Exclude the current incomplete week/month from analysis to avoid skewed trends"
            )
        else:
            st.markdown("""
            <div class="empty-state">
                <div class="empty-state-icon">📁</div>
                <div class="empty-state-title">Upload Data to Begin</div>
            </div>
            """, unsafe_allow_html=True)
            # Set default values before returning
            freq = "Monthly (default)"
            exclude_partial = True
            return
    
    # Load department mappings
    dept_mappings = load_department_mappings()
    
    # Get filtered data with loading indicator
    with st.spinner("📊 Loading data..."):
        if len(date_range) == 2:
            start_date, end_date = date_range
            data = db.get_filtered_data(
                start_date=start_date,
                end_date=end_date,
                departments=selected_depts if selected_depts else None
            )
        else:
            data = db.get_all_data()
    
    # Apply employee departments FIRST (authoritative source for employees)
    # This ensures the employee master file drives all employee department tagging
    data = apply_employee_departments(data)
    
    # Apply manual department mappings for non-employees (secondary/override)
    data = apply_department_mappings(data, dept_mappings)
    
    # Apply tool filter
    if selected_tool != 'All Tools' and not data.empty:
        data = data[data['tool_source'] == selected_tool]
    
    # Apply frequency normalization and partial period filtering
    if not data.empty:
        # Ensure date column is datetime
        data['date'] = pd.to_datetime(data['date'], errors='coerce')
        
        if freq.startswith("Weekly"):
            # Weekly view: normalize periods to ISO week starts
            # For OpenAI data (already weekly), keep as-is
            # For Blueflame data (monthly), allocate to weeks
            
            weekly_records = []
            for _, row in data.iterrows():
                if row['tool_source'] == 'BlueFlame AI':
                    # Allocate monthly data to weeks (even-by-day allocation)
                    month_start = pd.Timestamp(year=row['date'].year, month=row['date'].month, day=1)
                    month_end = month_start + pd.DateOffset(months=1) - pd.Timedelta(days=1)
                    
                    # Get all ISO week starts within this month (use set for performance)
                    current = month_start
                    weeks_in_month = set()
                    while current <= month_end:
                        week_start = current - pd.to_timedelta(current.weekday(), 'D')
                        weeks_in_month.add(week_start)
                        current += pd.Timedelta(days=7)
                    
                    # Convert back to sorted list for iteration
                    weeks_in_month = sorted(weeks_in_month)
                    
                    # Allocate usage evenly across weeks
                    if weeks_in_month:
                        usage_per_week = row['usage_count'] / len(weeks_in_month)
                        cost_per_week = row['cost_usd'] / len(weeks_in_month)
                        
                        for week_start in weeks_in_month:
                            weekly_row = row.copy()
                            weekly_row['period_start'] = week_start
                            weekly_row['usage_count'] = usage_per_week
                            weekly_row['cost_usd'] = cost_per_week
                            weekly_records.append(weekly_row)
                else:
                    # OpenAI data - already weekly, just ensure period_start is set to week start
                    week_start = row['date'] - pd.to_timedelta(row['date'].weekday(), 'D')
                    weekly_row = row.copy()
                    weekly_row['period_start'] = week_start
                    weekly_records.append(weekly_row)
            
            if weekly_records:
                data = pd.DataFrame(weekly_records)
                # Exclude partial current week if requested
                if exclude_partial:
                    today = pd.Timestamp.today().normalize()
                    current_week_start = today - pd.to_timedelta(today.weekday(), 'D')
                    data = data[data['period_start'] < current_week_start]
        else:
            # Monthly view: normalize periods to month starts
            # For Blueflame data (already monthly), keep as-is
            # For OpenAI data (weekly), prorate by day into calendar months
            
            monthly_records = []
            for _, row in data.iterrows():
                if row['tool_source'] in ['ChatGPT', 'OpenAI']:
                    # OpenAI weekly data - prorate into calendar months
                    # Get period_start, falling back to date column
                    period_start = pd.to_datetime(row.get('period_start', row['date']), errors='coerce')
                    if pd.isna(period_start):
                        period_start = row['date']
                    
                    # Estimate period_end as 6 days after period_start (7 day week)
                    period_end = period_start + pd.Timedelta(days=6)
                    
                    # Check if period spans multiple months
                    if period_start.month != period_end.month:
                        # Split across months
                        # Calculate days in each month
                        month1_end = pd.Timestamp(year=period_start.year, month=period_start.month, day=1) + pd.DateOffset(months=1) - pd.Timedelta(days=1)
                        days_in_month1 = (month1_end - period_start).days + 1
                        days_in_month2 = (period_end - month1_end).days
                        total_days = days_in_month1 + days_in_month2
                        
                        # Month 1 portion
                        month1_row = row.copy()
                        month1_row['period_start'] = pd.Timestamp(year=period_start.year, month=period_start.month, day=1)
                        month1_row['usage_count'] = row['usage_count'] * (days_in_month1 / total_days)
                        month1_row['cost_usd'] = row['cost_usd'] * (days_in_month1 / total_days)
                        monthly_records.append(month1_row)
                        
                        # Month 2 portion
                        month2_row = row.copy()
                        month2_row['period_start'] = pd.Timestamp(year=period_end.year, month=period_end.month, day=1)
                        month2_row['usage_count'] = row['usage_count'] * (days_in_month2 / total_days)
                        month2_row['cost_usd'] = row['cost_usd'] * (days_in_month2 / total_days)
                        monthly_records.append(month2_row)
                    else:
                        # Entire period in same month
                        monthly_row = row.copy()
                        monthly_row['period_start'] = pd.Timestamp(year=period_start.year, month=period_start.month, day=1)
                        monthly_records.append(monthly_row)
                else:
                    # Blueflame data - already monthly
                    monthly_row = row.copy()
                    monthly_row['period_start'] = pd.Timestamp(year=row['date'].year, month=row['date'].month, day=1)
                    monthly_records.append(monthly_row)
            
            if monthly_records:
                data = pd.DataFrame(monthly_records)
                # Exclude partial current month if requested
                if exclude_partial:
                    today = pd.Timestamp.today().normalize()
                    current_month_start = today.to_period('M').start_time
                    data = data[data['period_start'] < current_month_start]
    
    if data.empty:
        # Enhanced empty state for no filtered data
        st.markdown("""
        <div class="empty-state">
            <div class="empty-state-icon">🔍</div>
            <div class="empty-state-title">No Data Found</div>
            <div class="empty-state-text">
                No data matches your current filter selections.<br>
                Try adjusting your date range, tool, or department filters.
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Quick reset options
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Reset All Filters", type="secondary"):
                st.rerun()
        with col2:
            if st.button("📊 View All Data", type="primary"):
                st.session_state.clear()
                st.rerun()
        return
    
    # TAB 1: Executive Overview
    with tab1:
        # Clean header without emoji, with compact export menu
        col1, col2 = st.columns([4, 1])
        with col1:
            st.markdown('<h2 style="color: var(--text-primary); margin-bottom: 0;">Executive Summary</h2>', unsafe_allow_html=True)
            st.caption("Key performance metrics and trends")
        with col2:
            # Compact export menu in dropdown
            with st.expander("📥 Export", expanded=False):
                # PDF Export (HTML version)
                try:
                    html_content = generate_pdf_report_html(data, "AI Usage Executive Report")
                    st.download_button(
                        label="PDF Report",
                        data=html_content,
                        file_name=f"ai_usage_report_{datetime.now().strftime('%Y%m%d_%H%M')}.html",
                        mime="text/html",
                        use_container_width=True,
                        help="Download as HTML (print to PDF from browser)"
                    )
                except Exception as e:
                    st.error(f"Export error: {str(e)}")
                
                # Excel Export with pivot tables
                try:
                    excel_file = generate_excel_export(data, include_pivots=True)
                    st.download_button(
                        label="Excel Report",
                        data=excel_file,
                        file_name=f"ai_usage_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="Excel with pivot tables and summaries"
                    )
                except Exception as e:
                    st.error(f"Export error: {str(e)}")
        
        st.divider()
        
        # Weekly caption for Blueflame estimation
        if freq.startswith("Weekly"):
            st.caption("ⓘ Blueflame weekly values are estimated from monthly totals (even-by-day allocation).")
        
        # Calculate key usage metrics
        total_users = data['user_id'].nunique()
        total_usage = data['usage_count'].sum()
        avg_usage_per_user = total_usage / max(total_users, 1)
        
        # Enterprise License Notice
        st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">Enterprise License Notice</h3>', unsafe_allow_html=True)
        
        with st.expander("ℹ️ About Enterprise AI Tool Licensing", expanded=False):
            st.markdown("""
            **Why Costs Are Not Tracked in This Dashboard:**
            
            Enterprise AI tools (ChatGPT Enterprise, BlueFlame AI, etc.) use **license-based pricing models** 
            where costs are determined by:
            - Number of seats/licenses purchased (not message volume)
            - Contractual agreements negotiated at the enterprise level
            - Additional services and support packages
            
            ### This Dashboard Focuses On:
            
            1. **Usage Analytics** - How actively teams and individuals are using AI tools
            2. **Engagement Metrics** - Messages per user, active user trends, feature adoption
            3. **Department Insights** - Which departments are leveraging AI most effectively
            4. **Adoption Patterns** - Understanding which features resonate with users
            
            ### Why This Matters:
            
            - **ROI Understanding:** Usage volume and patterns indicate value delivered, regardless of pricing
            - **Training Opportunities:** Low usage may indicate need for training or awareness
            - **Feature Planning:** Popular features should be promoted and supported
            - **Resource Allocation:** High-usage departments may benefit from additional resources
            
            ### Key Metrics to Monitor:
            
            - **Active Users** - How many people are using the tools
            - **Messages per User** - Engagement depth and tool adoption
            - **Department Breakdown** - Understanding usage across the organization
            - **Feature Adoption** - Which capabilities are being utilized
            - **Trend Analysis** - Is usage growing, stable, or declining
            """)
        
        # Usage Analytics Overview
        st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">Usage Analytics Overview</h3>', unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "Total Active Users", 
                f"{total_users:,}",
                help="Unique users with AI tool activity across all platforms"
            )
            with st.expander("📊 Details"):
                st.write("**Calculation:**")
                st.code(f"COUNT(DISTINCT user_id) = {total_users:,}")
                st.write("Users with any message activity in the analyzed period")
        
        with col2:
            st.metric(
                "Total Messages",
                f"{total_usage:,}",
                help="Total interactions across all AI platforms"
            )
            with st.expander("📊 Details"):
                st.write("**Total Interactions:**")
                st.code(f"SUM(usage_count) = {total_usage:,}")
                if not data.empty and 'tool_source' in data.columns:
                    st.write("**By Provider:**")
                    provider_usage = data.groupby('tool_source')['usage_count'].sum().sort_values(ascending=False)
                    for provider, usage in provider_usage.items():
                        pct = (usage / total_usage * 100) if total_usage > 0 else 0
                        st.write(f"• {provider}: {usage:,} messages ({pct:.1f}%)")
        
        with col3:
            st.metric(
                "Messages per User", 
                f"{avg_usage_per_user:,.0f}",
                help="Average messages per user - key engagement indicator"
            )
            with st.expander("📊 Details"):
                st.write("**Engagement Calculation:**")
                st.code(f"{total_usage:,} messages ÷ {total_users} users = {avg_usage_per_user:,.0f} msgs/user")
                st.write("**Context:**")
                st.write(f"• Total messages: {total_usage:,}")
                st.write(f"• Total active users: {total_users:,}")
                
                # Show engagement by provider
                if not data.empty and 'tool_source' in data.columns:
                    st.write("**Messages per User by Provider:**")
                    provider_engagement = data.groupby('tool_source').agg({
                        'user_id': 'nunique',
                        'usage_count': 'sum'
                    })
                    provider_engagement['msgs_per_user'] = provider_engagement['usage_count'] / provider_engagement['user_id']
                    for provider, row in provider_engagement.iterrows():
                        st.write(f"• {provider}: {row['msgs_per_user']:,.0f} msgs/user")
        
        with col4:
            # Calculate active departments
            active_depts = data['department'].nunique() if 'department' in data.columns else 0
            st.metric(
                "Active Departments", 
                f"{active_depts}",
                help="Number of departments using AI tools"
            )
            with st.expander("📊 Details"):
                st.write("**Department Distribution:**")
                if not data.empty and 'department' in data.columns:
                    dept_counts = data['department'].value_counts().head(5)
                    st.write(f"**Top Departments by Activity:**")
                    for dept, count in dept_counts.items():
                        st.write(f"• {dept}: {count} records")
        
        st.divider()
        
        # Message Type Analytics Section
        st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">📊 Feature Adoption Analytics</h3>', unsafe_allow_html=True)
        
        st.markdown("""
        <div class="help-tooltip">
            💡 <strong>Message Type Breakdown:</strong> Understand which AI features your organization is using most.
            This helps identify adoption patterns and feature-specific training needs.
        </div>
        """, unsafe_allow_html=True)
        
        # Get organization-wide message breakdown
        org_breakdown = get_organization_message_breakdown(data)
        
        if org_breakdown:
            # Create visualization columns
            viz_col1, viz_col2 = st.columns([2, 1])
            
            with viz_col1:
                # Create pie chart for message type distribution
                breakdown_df = pd.DataFrame([
                    {'Feature': feature, 'Messages': count} 
                    for feature, count in org_breakdown.items()
                ])
                
                fig = px.pie(
                    breakdown_df,
                    values='Messages',
                    names='Feature',
                    title='Organization-Wide Feature Usage Distribution',
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Set3
                )
                
                fig.update_traces(
                    textposition='inside',
                    textinfo='percent+label',
                    hovertemplate='<b>%{label}</b><br>Messages: %{value:,}<br>Percentage: %{percent}<extra></extra>'
                )
                
                fig.update_layout(
                    showlegend=True,
                    legend=dict(
                        orientation='v',
                        yanchor='middle',
                        y=0.5,
                        xanchor='left',
                        x=1.05
                    ),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='white'),
                    height=400
                )
                
                st.plotly_chart(fig, use_container_width=True)
            
            with viz_col2:
                st.markdown("**Feature Usage Summary:**")
                
                total_messages = sum(org_breakdown.values())
                
                for feature, count in sorted(org_breakdown.items(), key=lambda x: x[1], reverse=True):
                    percentage = (count / total_messages * 100) if total_messages > 0 else 0
                    
                    # Create a styled metric card for each feature
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%); 
                                padding: 0.75rem; border-radius: 0.375rem; margin: 0.5rem 0;
                                border-left: 3px solid #667eea;">
                        <div style="font-size: 0.875rem; color: #64748b; margin-bottom: 0.25rem;">
                            {feature}
                        </div>
                        <div style="font-size: 1.25rem; font-weight: 600; color: #1e293b;">
                            {count:,}
                        </div>
                        <div style="font-size: 0.75rem; color: #94a3b8;">
                            {percentage:.1f}% of total
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Add feature descriptions
                st.markdown("---")
                st.markdown("**Feature Descriptions:**", help="Learn what each message type represents")
                
                with st.expander("ℹ️ What do these features mean?"):
                    st.markdown("""
                    **ChatGPT Messages:** Standard conversational AI interactions with ChatGPT
                    
                    **GPT Messages:** Custom GPT usage (specialized AI assistants tailored for specific tasks)
                    
                    **Tool Messages:** Advanced features like Code Interpreter, web browsing, and data analysis
                    
                    **Project Messages:** Messages within ChatGPT Projects (organized workspaces for collaboration)
                    
                    **BlueFlame Messages:** BlueFlame AI platform interactions (investment research and analysis)
                    """)
        else:
            st.info("📊 Message type analytics will appear here once data is loaded.")
        
        st.divider()
        
        # Data Quality & Validation Panel
        st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">Data Quality & Validation</h3>', unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        # Calculate completeness
        completeness = 100.0
        if not data.empty:
            required_cols = ['user_id', 'date', 'usage_count', 'cost_usd']
            for col in required_cols:
                if col in data.columns:
                    col_completeness = (data[col].notna().sum() / len(data)) * 100
                    completeness = min(completeness, col_completeness)
        
        # Get unique users count
        unique_users = data['user_id'].nunique() if not data.empty else 0
        
        # Calculate date coverage
        try:
            if not data.empty and 'date' in data.columns:
                valid_dates = pd.to_datetime(data['date'], errors='coerce').dropna()
                date_coverage = (valid_dates.max() - valid_dates.min()).days + 1
            else:
                date_coverage = 0
        except Exception as e:
            st.warning(f"⚠️ Date parsing issue: {str(e)}")
            date_coverage = 0
        
        with col1:
            quality_class = "quality-excellent" if completeness >= 95 else "quality-good" if completeness >= 80 else "quality-warning"
            st.markdown(f'<div class="{quality_class}" style="padding: 0.75rem; border-radius: 0.5rem; text-align: center;"><strong>{completeness:.1f}%</strong><br><small>Completeness</small></div>', unsafe_allow_html=True)
        
        with col2:
            st.markdown(f'<div class="quality-good" style="padding: 0.75rem; border-radius: 0.5rem; text-align: center;"><strong>{unique_users}</strong><br><small>Active Users</small></div>', unsafe_allow_html=True)
        
        with col3:
            st.markdown(f'<div class="quality-good" style="padding: 0.75rem; border-radius: 0.5rem; text-align: center;"><strong>{date_coverage}</strong><br><small>Days Coverage</small></div>', unsafe_allow_html=True)
        
        with col4:
            total_records = len(data)
            st.markdown(f'<div class="quality-good" style="padding: 0.75rem; border-radius: 0.5rem; text-align: center;"><strong>{total_records:,}</strong><br><small>Total Records</small></div>', unsafe_allow_html=True)
        
        # Data source breakdown
        if not data.empty and 'tool_source' in data.columns:
            st.markdown('<div style="margin-top: 1rem; padding: 1rem; background: #f8fafc; border-radius: 0.5rem; border: 1px solid #e2e8f0;">', unsafe_allow_html=True)
            st.write("**Data Sources:**")
            source_summary = data.groupby('tool_source').agg({
                'user_id': 'nunique',
                'usage_count': 'sum'
            }).reset_index()
            source_summary.columns = ['Provider', 'Users', 'Messages']
            
            total_messages = data['usage_count'].sum()
            for _, row in source_summary.iterrows():
                usage_pct = (row['Messages'] / total_messages * 100) if total_messages > 0 else 0
                st.write(f"**{row['Provider']}**: {row['Users']} users, {row['Messages']:,} messages ({usage_pct:.1f}% of total)")
            st.markdown('</div>', unsafe_allow_html=True)
        
        st.divider()
        
        # Month-over-Month Trends
        st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">Month-over-Month Trends</h3>', unsafe_allow_html=True)
        
        try:
            # Prepare monthly data
            monthly_data = data.copy()
            monthly_data['date'] = pd.to_datetime(monthly_data['date'], errors='coerce')
            monthly_data = monthly_data.dropna(subset=['date'])
            monthly_data['month'] = monthly_data['date'].dt.to_period('M').astype(str)
            
            # Calculate monthly metrics
            monthly_metrics = monthly_data.groupby('month').agg({
                'user_id': 'nunique',
                'usage_count': 'sum'
            }).reset_index()
            monthly_metrics.columns = ['Month', 'Active Users', 'Total Usage']
            
            # Calculate MoM changes
            if len(monthly_metrics) > 1:
                monthly_metrics['User Growth %'] = monthly_metrics['Active Users'].pct_change() * 100
                monthly_metrics['Usage Growth %'] = monthly_metrics['Total Usage'].pct_change() * 100
            
            col1, col2 = st.columns(2)
            
            with col1:
                # User adoption trend
                fig = px.line(
                    monthly_metrics, 
                    x='Month', 
                    y='Active Users',
                    title='Monthly Active Users Trend',
                    markers=True
                )
                fig.update_traces(line_color='#667eea', line_width=3, marker=dict(size=10))
                fig.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    hovermode='x unified'
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                # Usage trend
                fig = px.bar(
                    monthly_metrics, 
                    x='Month', 
                    y='Total Usage',
                    title='Monthly Message Volume Trend',
                    color='Total Usage',
                    color_continuous_scale='Blues'
                )
                fig.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    showlegend=False
                )
                st.plotly_chart(fig, use_container_width=True)
            
            # MoM Summary Table
            if len(monthly_metrics) > 1:
                st.markdown("**Month-over-Month Growth Summary:**")
                display_cols = ['Month', 'Active Users', 'User Growth %', 'Total Usage', 'Usage Growth %']
                st.dataframe(
                    monthly_metrics[display_cols].tail(6).style.format({
                        'Active Users': '{:,.0f}',
                        'User Growth %': '{:+.1f}%',
                        'Total Usage': '{:,.0f}',
                        'Usage Growth %': '{:+.1f}%'
                    }),
                    use_container_width=True,
                    hide_index=True
                )
        except Exception as e:
            st.warning(f"Unable to calculate monthly trends: {str(e)}")
        
        st.divider()
        
        # Department Performance Analytics
        st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">Department Performance</h3>', unsafe_allow_html=True)
        
        # Calculate comprehensive department statistics with message type breakdown
        dept_stats = data.groupby('department').agg({
            'user_id': 'nunique',
            'usage_count': 'sum'
        }).reset_index()
        dept_stats.columns = ['Department', 'Active Users', 'Total Usage']
        
        # Calculate message type breakdown for each department
        dept_message_breakdown = data.groupby(['department', 'feature_used'])['usage_count'].sum().reset_index()
        dept_message_pivot = dept_message_breakdown.pivot(index='department', columns='feature_used', values='usage_count').fillna(0)
        
        # Merge message type breakdown with dept_stats (single merge for better performance)
        dept_stats = dept_stats.merge(
            dept_message_pivot.reset_index().rename(columns={'department': 'Department'}),
            on='Department',
            how='left'
        ).fillna(0)
        
        # Calculate derived metrics
        total_usage_all = dept_stats['Total Usage'].sum()
        dept_stats['Usage Share %'] = (dept_stats['Total Usage'] / total_usage_all * 100).round(1)
        dept_stats['Avg Messages/User'] = (dept_stats['Total Usage'] / dept_stats['Active Users']).round(0)
        dept_stats = dept_stats.sort_values('Total Usage', ascending=False)
        
        # Add efficiency category
        conditions = [
            (dept_stats['Avg Messages/User'] > 5000),
            (dept_stats['Avg Messages/User'] > 2000),
            (dept_stats['Avg Messages/User'] <= 2000)
        ]
        values = ['High', 'Medium', 'Low']
        dept_stats['Efficiency'] = np.select(conditions, values, default='Medium')
        
        # Create tabs for different department views
        dept_tab1, dept_tab2, dept_tab3 = st.tabs([
            "📊 Department Comparison", 
            "👥 User Distribution", 
            "💡 Efficiency Analysis"
        ])
        
        with dept_tab1:
            st.markdown("### Department Usage Comparison")
            
            # Add interactive filtering for departments
            st.markdown("**🎯 Filter Chart Data**")
            filter_col1, filter_col2, filter_col3 = st.columns([2, 1, 1])
            
            with filter_col1:
                # Department filter - allow excluding specific departments
                all_departments = dept_stats['Department'].tolist()
                excluded_depts = st.multiselect(
                    "Exclude departments from chart (e.g., to remove outliers)",
                    all_departments,
                    help="Select departments to exclude from the visualization below"
                )
            
            with filter_col2:
                # Sort order
                sort_by = st.selectbox(
                    "Sort by:",
                    ["Total Usage", "Active Users", "Avg Messages/User"],
                    index=0
                )
            
            with filter_col3:
                # Number of departments to show
                if len(dept_stats) > 1:
                    num_depts = st.slider(
                        "Show top:",
                        min_value=1,
                        max_value=len(dept_stats),
                        value=min(7, len(dept_stats))
                    )
                else:
                    num_depts = len(dept_stats)
                    st.write(f"Showing: {num_depts} department" if num_depts == 1 else f"Showing: {num_depts} departments")
            
            # Apply filters
            dept_stats_filtered = dept_stats.copy()
            if excluded_depts:
                dept_stats_filtered = dept_stats_filtered[~dept_stats_filtered['Department'].isin(excluded_depts)]
            
            # Ensure num_depts doesn't exceed available filtered departments
            num_depts = min(num_depts, len(dept_stats_filtered))
            
            # Apply sorting
            sort_mapping = {
                "Total Usage": "Total Usage",
                "Active Users": "Active Users",
                "Avg Messages/User": "Avg Messages/User"
            }
            dept_stats_filtered = dept_stats_filtered.sort_values(by=sort_mapping[sort_by], ascending=False).head(num_depts)
            
            # ENHANCED VISUALIZATION: Create a dual-axis chart with stacked bars for message types
            fig = make_subplots(specs=[[{"secondary_y": True}]])
            
            # Get all available message types from the filtered data
            message_type_cols = [col for col in dept_stats_filtered.columns 
                               if col not in ['Department', 'Active Users', 'Total Usage', 'Usage Share %', 'Avg Messages/User', 'Efficiency']]
            
            # Define color scheme for message types
            message_type_colors = {
                'ChatGPT Messages': '#10a37f',
                'GPT Messages': '#0d8c6d',
                'Tool Messages': '#0e7a63',
                'Project Messages': '#0a6854',
                'BlueFlame Messages': '#4f46e5'
            }
            
            # Add stacked bars for each message type
            for msg_type in message_type_cols:
                fig.add_trace(
                    go.Bar(
                        x=dept_stats_filtered['Department'],
                        y=dept_stats_filtered[msg_type],
                        name=msg_type,
                        marker=dict(color=message_type_colors.get(msg_type, '#667eea')),
                        hovertemplate=f'<b>{msg_type}</b><br>%{{y:,}} messages<extra></extra>'
                    ),
                    secondary_y=False
                )
            
            # Add line for active users
            line = go.Scatter(
                x=dept_stats_filtered['Department'],
                y=dept_stats_filtered['Active Users'],
                name='Active Users',
                mode='markers+lines',
                marker=dict(size=12, symbol='circle', color='#FFA500'),
                line=dict(color='#FFA500', width=3),
                hovertemplate='<b>%{x}</b><br>Active Users: %{y}<extra></extra>',
                yaxis='y2'
            )
            
            # Add the line trace
            fig.add_trace(line, secondary_y=True)
            
            # Update layout with improved styling
            fig.update_layout(
                title=f"Department Usage Comparison - Message Type Breakdown ({len(dept_stats_filtered)} of {len(dept_stats)} depts)",
                xaxis=dict(title='Department', tickangle=-45, tickfont=dict(size=11)),
                yaxis=dict(title='Total Messages', gridcolor='rgba(255,255,255,0.1)'),
                yaxis2=dict(
                    title='Active Users', 
                    gridcolor='rgba(255,255,255,0)', 
                    range=[0, max(dept_stats_filtered['Active Users'])*1.2] if not dept_stats_filtered.empty else [0, 1]
                ),
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
                barmode='stack',  # Changed to stack for message types
                height=500,
                margin=dict(t=80, b=100),
                hovermode='closest',
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white')
            )
            
            # Add total messages and avg messages per user annotations
            # Place 'Total Messages' at the top, 'Messages per User' inside bars
            for i, row in dept_stats_filtered.iterrows():
                # Total messages at top of stacked bar
                fig.add_annotation(
                    x=row['Department'],
                    y=row['Total Usage'],
                    text=f"{row['Total Usage']:,}",
                    showarrow=False,
                    font=dict(size=10, color="white", weight='bold'),
                    xanchor='center',
                    yanchor='bottom',
                    yshift=5
                )
                
                # Messages per user inside the bar (at midpoint)
                fig.add_annotation(
                    x=row['Department'],
                    y=row['Total Usage'] / 2,  # Middle of the bar
                    text=f"{row['Avg Messages/User']:,.0f}/user",
                    showarrow=False,
                    font=dict(size=9, color="rgba(255,255,255,0.8)"),
                    xanchor='center',
                    yanchor='middle'
                )
            
            # Display chart
            st.plotly_chart(fig, use_container_width=True, key='dept_comparison_chart')
            
            # Department drilldown selector (use all departments, not just filtered ones)
            st.markdown("---")
            st.markdown("### 👥 Department User Drilldown")
            st.caption("Select a department to view detailed user-level statistics")
            
            # Create department selector with all departments
            dept_selector_col, spacer_col = st.columns([3, 2])
            with dept_selector_col:
                selected_dept = st.selectbox(
                    "Choose a department:",
                    options=['[Select a department]'] + dept_stats['Department'].tolist(),
                    key='dept_drilldown_selector'
                )
            
            # Show user details if a department is selected
            if selected_dept and selected_dept != '[Select a department]':
                st.markdown(f"#### 📊 Users in {selected_dept}")
                
                # Get users for this department
                dept_users = data[data['department'] == selected_dept].copy()
                
                if not dept_users.empty:
                    # Aggregate by user
                    user_stats = dept_users.groupby(['email', 'user_name']).agg({
                        'usage_count': 'sum',
                        'date': ['min', 'max']
                    }).reset_index()
                    
                    # Flatten multi-level columns
                    user_stats.columns = ['email', 'user_name', 'total_messages', 'first_active', 'last_active']
                    
                    # Get message type breakdown for each user
                    for msg_type in message_type_cols:
                        user_msg_type = dept_users[dept_users['feature_used'] == msg_type].groupby('email')['usage_count'].sum()
                        user_stats[msg_type] = user_stats['email'].map(user_msg_type).fillna(0).astype(int)
                    
                    # Sort by total messages
                    user_stats = user_stats.sort_values('total_messages', ascending=False)
                    
                    # Summary metrics for selected department
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Users", len(user_stats))
                    with col2:
                        st.metric("Total Messages", f"{user_stats['total_messages'].sum():,}")
                    with col3:
                        st.metric("Avg Messages/User", f"{user_stats['total_messages'].mean():,.0f}")
                    
                    st.markdown("**User Details:**")
                    
                    # Display user table
                    st.dataframe(
                        user_stats,
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            'email': st.column_config.TextColumn('Email', width='medium'),
                            'user_name': st.column_config.TextColumn('Name', width='medium'),
                            'total_messages': st.column_config.NumberColumn('Total Messages', format='%d'),
                            'first_active': st.column_config.DateColumn('First Active'),
                            'last_active': st.column_config.DateColumn('Last Active'),
                            **{msg_type: st.column_config.NumberColumn(msg_type, format='%d') for msg_type in message_type_cols if msg_type in user_stats.columns}
                        }
                    )
                    
                    # Add download button for this department's users
                    csv = user_stats.to_csv(index=False)
                    st.download_button(
                        label=f"📥 Download {selected_dept} Users CSV",
                        data=csv,
                        file_name=f"{selected_dept.lower().replace(' ', '_')}_users.csv",
                        mime="text/csv"
                    )
                else:
                    st.info(f"No user data available for {selected_dept}")
            
            
            # USING THE SPACE BELOW: Add a detailed data table
            st.markdown("### Department Details")
            
            # Enhanced table with efficiency indicators
            table_df = dept_stats_filtered.copy()
            
            # Format columns for display
            table_df['Total Usage Formatted'] = table_df['Total Usage'].apply(lambda x: f"{x:,}")
            table_df['Avg Messages/User Formatted'] = table_df['Avg Messages/User'].apply(lambda x: f"{x:,}")
            table_df['Usage Share % Formatted'] = table_df['Usage Share %'].apply(lambda x: f"{x:.1f}%")
            
            # Create display dataframe
            display_df = table_df[[
                'Department', 'Total Usage Formatted', 'Active Users', 
                'Avg Messages/User Formatted', 'Usage Share % Formatted', 'Efficiency'
            ]].copy()
            display_df.columns = [
                'Department', 'Total Messages', 'Active Users',
                'Messages/User', '% of Total', 'Efficiency'
            ]
            
            # Display styled table
            st.markdown('<div class="dataframe-container">', unsafe_allow_html=True)
            
            # Function to highlight efficiency
            def highlight_efficiency(row):
                if row['Efficiency'] == 'High':
                    return ['background-color: rgba(21, 128, 61, 0.3); color: #ffffff;'] * len(row)
                elif row['Efficiency'] == 'Medium':
                    return ['background-color: rgba(202, 138, 4, 0.3); color: #ffffff;'] * len(row)
                else:
                    return ['background-color: rgba(220, 38, 38, 0.3); color: #ffffff;'] * len(row)
            
            st.dataframe(
                display_df,
                use_container_width=True,
                hide_index=True
            )
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Add expandable message type breakdown per department
            st.markdown("---")
            st.markdown("### 📊 Feature Usage by Department")
            st.caption("Click on a department to see detailed feature usage breakdown")
            
            for _, dept_row in dept_stats_filtered.iterrows():
                dept_name = dept_row['Department']
                dept_breakdown = get_department_message_breakdown(data, dept_name)
                
                if dept_breakdown:
                    with st.expander(f"🏢 {dept_name} - Feature Breakdown ({dept_row['Total Usage']:,} total messages)"):
                        # Create columns for metrics
                        breakdown_col1, breakdown_col2 = st.columns([2, 1])
                        
                        with breakdown_col1:
                            # Create a bar chart for this department's feature usage
                            breakdown_df = pd.DataFrame([
                                {'Feature': feature, 'Messages': count} 
                                for feature, count in dept_breakdown.items()
                            ]).sort_values('Messages', ascending=False)
                            
                            fig_dept = px.bar(
                                breakdown_df,
                                x='Feature',
                                y='Messages',
                                title=f'{dept_name} - Feature Usage',
                                text='Messages',
                                color='Messages',
                                color_continuous_scale='Blues'
                            )
                            
                            fig_dept.update_traces(
                                texttemplate='%{text:,}',
                                textposition='outside'
                            )
                            
                            fig_dept.update_layout(
                                xaxis_title='Feature Type',
                                yaxis_title='Message Count',
                                plot_bgcolor='rgba(0,0,0,0)',
                                paper_bgcolor='rgba(0,0,0,0)',
                                font=dict(color='white'),
                                height=300,
                                showlegend=False
                            )
                            
                            st.plotly_chart(fig_dept, use_container_width=True)
                        
                        with breakdown_col2:
                            st.markdown("**Feature Details:**")
                            total_dept_messages = sum(dept_breakdown.values())
                            
                            for feature, count in sorted(dept_breakdown.items(), key=lambda x: x[1], reverse=True):
                                percentage = (count / total_dept_messages * 100) if total_dept_messages > 0 else 0
                                st.markdown(f"""
                                <div style="background: var(--card-bg); padding: 0.5rem; 
                                            border-radius: 0.25rem; margin: 0.25rem 0;">
                                    <div style="font-size: 0.75rem; color: #94a3b8;">{feature}</div>
                                    <div style="font-size: 1rem; font-weight: 600;">{count:,}</div>
                                    <div style="font-size: 0.7rem; color: #64748b;">{percentage:.1f}% of dept total</div>
                                </div>
                                """, unsafe_allow_html=True)
                            
                            # Add department-specific insights
                            st.markdown("---")
                            st.markdown(f"**Department Stats:**")
                            st.write(f"• Active Users: {dept_row['Active Users']}")
                            st.write(f"• Messages/User: {dept_row['Avg Messages/User']:,.0f}")
            
            # Add key insights section
            st.markdown("---")
            st.markdown("### Key Insights")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Most efficient department
                most_efficient = dept_stats.loc[dept_stats['Avg Messages/User'].idxmax()]
                st.markdown(f"""
                <div class="insight-card insight-success">
                    <h4>Most Efficient Department</h4>
                    <p><strong>{most_efficient['Department']}</strong></p>
                    <p>{most_efficient['Avg Messages/User']:,.0f} messages per user</p>
                    <p>Total messages: {most_efficient['Total Usage']:,.0f} | Users: {most_efficient['Active Users']}</p>
                </div>
                """, unsafe_allow_html=True)
                
                # Department with most users
                most_users = dept_stats.loc[dept_stats['Active Users'].idxmax()]
                st.markdown(f"""
                <div class="insight-card insight-info">
                    <h4>Highest User Adoption</h4>
                    <p><strong>{most_users['Department']}</strong></p>
                    <p>{most_users['Active Users']} active users</p>
                    <p>Average: {most_users['Avg Messages/User']:,.0f} messages per user</p>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                # Usage insights
                highest_share = dept_stats.loc[dept_stats['Usage Share %'].idxmax()]
                st.markdown(f"""
                <div class="insight-card insight-warning">
                    <h4>Largest Usage Share</h4>
                    <p><strong>{highest_share['Department']}</strong></p>
                    <p>{highest_share['Usage Share %']:.1f}% of total usage</p>
                    <p>{highest_share['Total Usage']:,} messages ({highest_share['Active Users']} users)</p>
                </div>
                """, unsafe_allow_html=True)
                
                # Efficiency distribution
                efficiency_counts = dept_stats['Efficiency'].value_counts()
                st.markdown(f"""
                <div class="insight-card insight-success">
                    <h4>Efficiency Distribution</h4>
                    <div class="metric-row">
                        <span class="metric-label">High Efficiency Departments:</span>
                        <span class="metric-value">{efficiency_counts.get('High', 0)}</span>
                    </div>
                    <div class="metric-row">
                        <span class="metric-label">Medium Efficiency Departments:</span>
                        <span class="metric-value">{efficiency_counts.get('Medium', 0)}</span>
                    </div>
                    <div class="metric-row">
                        <span class="metric-label">Low Efficiency Departments:</span>
                        <span class="metric-value">{efficiency_counts.get('Low', 0)}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        with dept_tab2:
            st.markdown("### User Distribution by Department")
            
            # Create a horizontal bar chart that shows users per department
            fig = px.bar(
                dept_stats.sort_values('Active Users', ascending=True),
                y='Department',
                x='Active Users',
                orientation='h',
                text='Active Users',
                title='Active Users by Department',
                color='Avg Messages/User',
                color_continuous_scale='Blues',
                hover_data=['Total Usage', 'Avg Messages/User']
            )
            
            fig.update_layout(
                xaxis_title="Active Users",
                yaxis_title="Department",
                height=600,
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white')
            )
            
            fig.update_traces(
                textposition='outside',
                textfont=dict(size=12)
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # Add user breakdown cards
            st.markdown("### User Activity Breakdown")
            
            # Create a grid of cards for departments with user breakdown
            cols = st.columns(3)
            for i, (_, row) in enumerate(dept_stats.iterrows()):
                with cols[i % 3]:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h4>{row['Department']}</h4>
                        <div class="metric-row">
                            <span class="metric-label">Active Users:</span>
                            <span class="metric-value">{row['Active Users']}</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Total Messages:</span>
                            <span class="metric-value">{row['Total Usage']:,.0f}</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Messages per User:</span>
                            <span class="metric-value">{row['Avg Messages/User']:,.0f}</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Efficiency:</span>
                            <span class="metric-value">
                                {row['Efficiency']}
                                <span class="efficiency-badge {row['Efficiency'].lower()}">{row['Efficiency']}</span>
                            </span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
        
        with dept_tab3:
            st.markdown("### Department Efficiency Analysis")
            
            # Create a scatter plot of messages vs users
            fig = px.scatter(
                dept_stats,
                x='Active Users',
                y='Total Usage',
                size='Avg Messages/User',
                color='Efficiency',
                hover_name='Department',
                text='Department',
                color_discrete_map={
                    'High': '#15803d',
                    'Medium': '#ca8a04',
                    'Low': '#dc2626'
                },
                title='Department Efficiency Matrix (Messages vs Users)',
                size_max=50,
                labels={
                    'Total Usage': 'Total Messages',
                    'Active Users': 'Active Users',
                    'Efficiency': 'Efficiency'
                }
            )
            
            # Add reference lines for average efficiency
            avg_msg_per_user = dept_stats['Avg Messages/User'].mean()
            max_users = dept_stats['Active Users'].max()
            max_messages = dept_stats['Total Usage'].max()
            
            # Add diagonal reference lines for different efficiency levels
            for efficiency, color in zip([avg_msg_per_user/2, avg_msg_per_user, avg_msg_per_user*2], 
                                       ['rgba(220, 38, 38, 0.4)', 'rgba(202, 138, 4, 0.4)', 'rgba(21, 128, 61, 0.4)']):
                x_vals = list(range(1, int(max_users * 1.1)))
                y_vals = [x * efficiency for x in x_vals]
                
                fig.add_trace(
                    go.Scatter(
                        x=x_vals,
                        y=y_vals,
                        mode='lines',
                        line=dict(color=color, width=1, dash='dot'),
                        name=f'{int(efficiency):,} msg/user',
                        hoverinfo='name'
                    )
                )
            
            # Add annotations for diagonal reference lines
            fig.add_annotation(
                x=max_users * 0.7,
                y=max_users * 0.7 * avg_msg_per_user,
                text=f"Avg: {int(avg_msg_per_user):,} msg/user",
                showarrow=False,
                font=dict(color='rgba(202, 138, 4, 1)')
            )
            
            # Update layout
            fig.update_layout(
                height=600,
                xaxis=dict(title='Active Users'),
                yaxis=dict(title='Total Messages'),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white')
            )
            
            # Update traces
            fig.update_traces(
                textposition='top center',
                textfont=dict(size=10)
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            # Add efficiency distribution
            st.markdown("### Efficiency Distribution")
            
            # Create columns for metrics and chart
            col1, col2 = st.columns([1, 2])
            
            with col1:
                # Summary metrics
                st.markdown(f"""
                <div class="metric-card">
                    <h4>Efficiency Metrics</h4>
                    <div class="metric-row">
                        <span class="metric-label">Average Efficiency:</span>
                        <span class="metric-value">{dept_stats['Avg Messages/User'].mean():,.0f} msgs/user</span>
                    </div>
                    <div class="metric-row">
                        <span class="metric-label">Highest Efficiency:</span>
                        <span class="metric-value">{dept_stats['Avg Messages/User'].max():,.0f} msgs/user</span>
                    </div>
                    <div class="metric-row">
                        <span class="metric-label">Lowest Efficiency:</span>
                        <span class="metric-value">{dept_stats['Avg Messages/User'].min():,.0f} msgs/user</span>
                    </div>
                    <div class="metric-row">
                        <span class="metric-label">Median Efficiency:</span>
                        <span class="metric-value">{dept_stats['Avg Messages/User'].median():,.0f} msgs/user</span>
                    </div>
                    <div class="metric-row">
                        <span class="metric-label">Total Messages:</span>
                        <span class="metric-value">{dept_stats['Total Usage'].sum():,}</span>
                    </div>
                    <div class="metric-row">
                        <span class="metric-label">Total Users:</span>
                        <span class="metric-value">{dept_stats['Active Users'].sum()}</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                # Create histogram of messages per user
                hist_fig = px.histogram(
                    dept_stats,
                    x='Avg Messages/User',
                    nbins=10,
                    color='Efficiency',
                    color_discrete_map={
                        'High': '#15803d',
                        'Medium': '#ca8a04',
                        'Low': '#dc2626'
                    },
                    title='Distribution of Efficiency (Messages per User)',
                    labels={'Avg Messages/User': 'Messages per User'}
                )
                
                hist_fig.update_layout(
                    bargap=0.1,
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='white'),
                    showlegend=True
                )
                
                # Add mean line
                hist_fig.add_vline(
                    x=dept_stats['Avg Messages/User'].mean(),
                    line_dash="solid",
                    line_color="white",
                    annotation_text="Mean",
                    annotation_position="top right"
                )
                
                st.plotly_chart(hist_fig, use_container_width=True)
        
        st.divider()
        
        # Original insights below (kept for backwards compatibility)
        st.markdown("**Summary Insights**")
        insight_cols = st.columns(3)
        
        with insight_cols[0]:
            highest_volume = dept_stats.loc[dept_stats['Total Usage'].idxmax()]
            st.markdown(f"""
            <div class="insight-card insight-success">
                <h4 style="margin: 0 0 0.5rem 0; color: var(--success-text);">📊 Highest Volume</h4>
                <p style="margin: 0; font-weight: 600; color: var(--success-text);">{highest_volume['Department']}</p>
                <p style="margin: 0; font-size: 0.875rem; color: var(--success-text);">{highest_volume['Total Usage']:,} total messages</p>
            </div>
            """, unsafe_allow_html=True)
        
        with insight_cols[1]:
            most_active = dept_stats.loc[dept_stats['Active Users'].idxmax()]
            st.markdown(f"""
            <div class="insight-card insight-info">
                <h4 style="margin: 0 0 0.5rem 0; color: var(--info-text);">👥 Highest Adoption</h4>
                <p style="margin: 0; font-weight: 600; color: var(--info-text);">{most_active['Department']}</p>
                <p style="margin: 0; font-size: 0.875rem; color: var(--info-text);">{most_active['Active Users']} active users</p>
            </div>
            """, unsafe_allow_html=True)
        
        with insight_cols[2]:
            most_engaged = dept_stats.loc[dept_stats['Avg Messages/User'].idxmax()]
            st.markdown(f"""
            <div class="insight-card insight-warning">
                <h4 style="margin: 0 0 0.5rem 0; color: var(--warning-text);">🚀 Most Engaged</h4>
                <p style="margin: 0; font-weight: 600; color: var(--warning-text);">{most_engaged['Department']}</p>
                <p style="margin: 0; font-size: 0.875rem; color: var(--warning-text);">{most_engaged['Avg Messages/User']:,.0f} avg messages/user</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.divider()
    
    # TAB 2: Tool Comparison
    with tab2:
        display_tool_comparison(data)
    
    # TAB: OpenAI Analytics
    with tab_openai:
        st.header("🤖 OpenAI Usage Deep Dive")
        
        st.markdown("""
        <div class="help-tooltip">
            💡 <strong>OpenAI-Specific Analytics:</strong> Deep dive into ChatGPT Enterprise usage patterns,
            feature adoption, and user engagement metrics. This view shows ONLY OpenAI data.
        </div>
        """, unsafe_allow_html=True)
        
        # Filter to OpenAI data only
        openai_data = get_openai_data(data)
        
        if openai_data.empty:
            st.markdown("""
            <div class="empty-state">
                <div class="empty-state-icon">🤖</div>
                <div class="empty-state-title">No OpenAI Data Available</div>
                <div class="empty-state-text">
                    Upload OpenAI ChatGPT Enterprise exports to see detailed analytics here.
                </div>
            </div>
            """, unsafe_allow_html=True)
        else:
            # Executive Summary Metrics
            st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">📊 Executive Summary</h3>', unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                total_active = openai_data['user_id'].nunique()
                st.metric(
                    "Total Active Users",
                    f"{total_active:,}",
                    help="Unique users with OpenAI ChatGPT activity"
                )
                with st.expander("📊 Details"):
                    st.write("**Calculation:**")
                    st.code(f"COUNT(DISTINCT user_id) = {total_active:,}")
                    st.write(f"Users with any ChatGPT message activity in the selected period")
            
            with col2:
                duau = calculate_duau(openai_data)
                st.metric(
                    "Daily Unique Active Users",
                    f"{duau:,.0f}",
                    help="Average unique users active per day"
                )
                with st.expander("📊 Details"):
                    st.write("**Calculation:**")
                    st.code(f"AVG(daily unique users) = {duau:,.0f}")
                    st.write("Average number of unique users active on any given day")
            
            with col3:
                total_messages = openai_data['usage_count'].sum()
                avg_per_user = total_messages / max(total_active, 1)
                st.metric(
                    "Messages per User",
                    f"{avg_per_user:,.0f}",
                    help="Average messages per user"
                )
                with st.expander("📊 Details"):
                    st.write("**Calculation:**")
                    st.code(f"{total_messages:,} messages ÷ {total_active} users = {avg_per_user:,.0f}")
                    st.write(f"Total messages: {total_messages:,}")
            
            with col4:
                # Calculate days active per month
                days_active_df = calculate_days_active_per_month(openai_data)
                avg_days = days_active_df['avg_days_per_month'].mean() if not days_active_df.empty else 0
                st.metric(
                    "Avg Days Active/Month",
                    f"{avg_days:.1f}",
                    help="Average days per month users are active"
                )
                with st.expander("📊 Details"):
                    st.write("**Calculation:**")
                    st.code(f"AVG(days active per user per month) = {avg_days:.1f}")
                    st.write("Average number of distinct days each user was active in a month")
            
            st.divider()
            
            # Message Type Breakdown
            st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">💬 Message Type Distribution</h3>', unsafe_allow_html=True)
            
            # Get message breakdown
            message_breakdown = openai_data.groupby('feature_used')['usage_count'].sum().reset_index()
            message_breakdown.columns = ['Feature', 'Messages']
            message_breakdown = message_breakdown.sort_values('Messages', ascending=False)
            
            col1, col2 = st.columns([3, 2])
            
            with col1:
                # Create donut chart
                fig = px.pie(
                    message_breakdown,
                    values='Messages',
                    names='Feature',
                    title='OpenAI Feature Usage Distribution',
                    hole=0.4,
                    color_discrete_sequence=px.colors.qualitative.Set2
                )
                
                fig.update_traces(
                    textposition='inside',
                    textinfo='percent+label',
                    hovertemplate='<b>%{label}</b><br>Messages: %{value:,}<br>Percentage: %{percent}<extra></extra>'
                )
                
                fig.update_layout(
                    showlegend=True,
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='white'),
                    height=400
                )
                
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                st.markdown("**Feature Breakdown:**")
                
                total_msgs = message_breakdown['Messages'].sum()
                
                for _, row in message_breakdown.iterrows():
                    pct = (row['Messages'] / total_msgs * 100) if total_msgs > 0 else 0
                    
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%); 
                                padding: 0.75rem; border-radius: 0.375rem; margin: 0.5rem 0;
                                border-left: 3px solid #667eea;">
                        <div style="font-size: 0.875rem; color: #64748b; margin-bottom: 0.25rem;">
                            {row['Feature']}
                        </div>
                        <div style="font-size: 1.25rem; font-weight: 600; color: #1e293b;">
                            {row['Messages']:,}
                        </div>
                        <div style="font-size: 0.75rem; color: #94a3b8;">
                            {pct:.1f}% of total
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Feature descriptions
                with st.expander("ℹ️ What are these features?"):
                    st.markdown("""
                    **ChatGPT Messages:** Standard conversational interactions with ChatGPT
                    
                    **GPT Messages:** Custom GPT usage (specialized AI assistants)
                    
                    **Tool Messages:** Advanced features - Code Interpreter, web browsing, DALL-E, data analysis
                    
                    **Project Messages:** Messages within ChatGPT Projects (organized collaborative workspaces)
                    """)
            
            st.divider()
            
            # User Activity Tiers
            st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">👥 User Engagement Analysis</h3>', unsafe_allow_html=True)
            
            user_tiers = get_user_activity_tiers(openai_data)
            
            if not user_tiers.empty:
                # Tier distribution
                tier_counts = user_tiers['activity_tier'].value_counts()
                
                # Initialize session state for tier selection
                if 'selected_tier' not in st.session_state:
                    st.session_state.selected_tier = None
                
                col1, col2, col3 = st.columns([2, 2, 2])
                
                with col1:
                    # Tier distribution pie chart
                    fig_tiers = px.pie(
                        values=tier_counts.values,
                        names=tier_counts.index,
                        title='User Engagement Tiers',
                        color=tier_counts.index,
                        color_discrete_map={
                            'Heavy': '#15803d',
                            'Moderate': '#ca8a04',
                            'Light': '#dc2626',
                            'Inactive': '#94a3b8'
                        }
                    )
                    
                    fig_tiers.update_traces(
                        textinfo='percent+label',
                        hovertemplate='<b>%{label}</b><br>Users: %{value}<br>Percentage: %{percent}<extra></extra>'
                    )
                    
                    fig_tiers.update_layout(
                        showlegend=True,
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='white'),
                        height=350
                    )
                    
                    st.plotly_chart(fig_tiers, use_container_width=True)
                
                with col2:
                    st.markdown("**Tier Definitions:**")
                    
                    # Heavy Users - clickable card
                    if st.button("🔥 Heavy Users", key="heavy_btn", use_container_width=True):
                        st.session_state.selected_tier = 'Heavy'
                    
                    st.markdown(f"""
                    <div class="metric-card">
                        <div style="font-size: 0.9rem; color: #64748b; margin-bottom: 0.5rem;">
                            Top 20% by message volume
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Count:</span>
                            <span class="metric-value">{tier_counts.get('Heavy', 0)} users</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Avg Messages:</span>
                            <span class="metric-value">{user_tiers[user_tiers['activity_tier']=='Heavy']['total_messages'].mean() if len(user_tiers[user_tiers['activity_tier']=='Heavy']) > 0 else 0:,.0f}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Moderate Users - clickable card
                    if st.button("📊 Moderate Users", key="moderate_btn", use_container_width=True):
                        st.session_state.selected_tier = 'Moderate'
                    
                    st.markdown(f"""
                    <div class="metric-card" style="margin-top: 1rem;">
                        <div style="font-size: 0.9rem; color: #64748b; margin-bottom: 0.5rem;">
                            50th-80th percentile
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Count:</span>
                            <span class="metric-value">{tier_counts.get('Moderate', 0)} users</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Avg Messages:</span>
                            <span class="metric-value">{user_tiers[user_tiers['activity_tier']=='Moderate']['total_messages'].mean() if len(user_tiers[user_tiers['activity_tier']=='Moderate']) > 0 else 0:,.0f}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col3:
                    # Light Users - clickable card
                    if st.button("💡 Light Users", key="light_btn", use_container_width=True):
                        st.session_state.selected_tier = 'Light'
                    
                    st.markdown(f"""
                    <div class="metric-card">
                        <div style="font-size: 0.9rem; color: #64748b; margin-bottom: 0.5rem;">
                            20th-50th percentile
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Count:</span>
                            <span class="metric-value">{tier_counts.get('Light', 0)} users</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Avg Messages:</span>
                            <span class="metric-value">{user_tiers[user_tiers['activity_tier']=='Light']['total_messages'].mean() if len(user_tiers[user_tiers['activity_tier']=='Light']) > 0 else 0:,.0f}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Inactive Users - clickable card with special highlight
                    if st.button("😴 Inactive Users", key="inactive_btn", use_container_width=True):
                        st.session_state.selected_tier = 'Inactive'
                    
                    st.markdown(f"""
                    <div class="metric-card" style="margin-top: 1rem; border: 2px solid #f59e0b;">
                        <div style="font-size: 0.9rem; color: #64748b; margin-bottom: 0.5rem;">
                            Bottom 20% by usage - <strong style="color: #f59e0b;">Priority Follow-up</strong>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Count:</span>
                            <span class="metric-value">{tier_counts.get('Inactive', 0)} users</span>
                        </div>
                        <div class="metric-row">
                            <span class="metric-label">Avg Messages:</span>
                            <span class="metric-value">{user_tiers[user_tiers['activity_tier']=='Inactive']['total_messages'].mean() if len(user_tiers[user_tiers['activity_tier']=='Inactive']) > 0 else 0:,.0f}</span>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
            
            # User Drilldown Section
            if st.session_state.selected_tier:
                st.divider()
                
                tier_data = user_tiers[user_tiers['activity_tier'] == st.session_state.selected_tier].copy()
                
                # Tier color mapping
                tier_colors = {
                    'Heavy': '#15803d',
                    'Moderate': '#ca8a04',
                    'Light': '#dc2626',
                    'Inactive': '#94a3b8'
                }
                
                tier_icons = {
                    'Heavy': '🔥',
                    'Moderate': '📊',
                    'Light': '💡',
                    'Inactive': '😴'
                }
                
                col_header, col_close = st.columns([6, 1])
                with col_header:
                    st.markdown(f'<h3 style="color: {tier_colors[st.session_state.selected_tier]};">{tier_icons[st.session_state.selected_tier]} {st.session_state.selected_tier} Users - Detailed View</h3>', unsafe_allow_html=True)
                with col_close:
                    if st.button("✖ Close", key="close_drilldown"):
                        st.session_state.selected_tier = None
                        st.rerun()
                
                if st.session_state.selected_tier == 'Inactive':
                    st.warning("⚠️ **Priority Follow-up Demographic**: These users show minimal engagement and may benefit from support or training.")
                
                # Display user details table
                st.markdown(f"**{len(tier_data)} users in this tier**")
                
                # Prepare display dataframe
                display_cols = ['user_name', 'email', 'department', 'total_messages']
                
                # Add activity date columns if they exist
                if 'last_day_active' in tier_data.columns:
                    display_cols.append('last_day_active')
                if 'last_day_active_in_period' in tier_data.columns:
                    display_cols.append('last_day_active_in_period')
                if 'first_day_active_in_period' in tier_data.columns:
                    display_cols.append('first_day_active_in_period')
                
                # Add feature breakdown columns
                if 'ChatGPT Messages' in tier_data.columns:
                    display_cols.append('ChatGPT Messages')
                if 'Tool Messages' in tier_data.columns:
                    display_cols.append('Tool Messages')
                if 'Project Messages' in tier_data.columns:
                    display_cols.append('Project Messages')
                
                # Filter to only columns that exist
                display_cols = [col for col in display_cols if col in tier_data.columns]
                
                display_df = tier_data[display_cols].sort_values('total_messages', ascending=False).reset_index(drop=True)
                
                # Rename columns for better display
                column_renames = {
                    'user_name': 'Name',
                    'email': 'Email',
                    'department': 'Department',
                    'total_messages': 'Total Messages',
                    'last_day_active': 'Last Sign-In',
                    'last_day_active_in_period': 'Last Active in Period',
                    'first_day_active_in_period': 'First Active in Period',
                    'ChatGPT Messages': 'ChatGPT Msgs',
                    'Tool Messages': 'Tool Msgs',
                    'Project Messages': 'Project Msgs'
                }
                
                display_df.rename(columns=column_renames, inplace=True)
                
                # Display the dataframe with formatting
                st.dataframe(
                    display_df,
                    use_container_width=True,
                    height=400,
                    hide_index=True
                )
                
                # Add export button
                csv = display_df.to_csv(index=False)
                st.download_button(
                    label=f"📥 Download {st.session_state.selected_tier} Users CSV",
                    data=csv,
                    file_name=f"{st.session_state.selected_tier.lower()}_users.csv",
                    mime="text/csv"
                )
            
            st.divider()
            
            # Weekly Trends
            st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">📅 Weekly Activity Trends</h3>', unsafe_allow_html=True)
            st.caption("📊 OpenAI weekly data only (no Blueflame data included)")
            
            weekly_data = calculate_weekly_trends(openai_data)
            
            if not weekly_data.empty:
                col1, col2 = st.columns(2)
                
                with col1:
                    # Active users trend
                    fig_users = px.line(
                        weekly_data,
                        x='week_display',
                        y='active_users',
                        title='Weekly Active Users',
                        markers=True,
                        labels={'week_display': 'Week Starting', 'active_users': 'Active Users'}
                    )
                    
                    fig_users.update_traces(
                        line_color='#667eea',
                        line_width=3,
                        marker=dict(size=8),
                        hovertemplate='%{y}<extra></extra>'  # Show only value; <extra></extra> removes trace name box
                    )
                    
                    fig_users.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='white'),
                        hovermode='x unified',
                        height=350,
                        xaxis=dict(
                            tickangle=-45,
                            tickmode='auto',
                            nticks=10
                        )
                    )
                    
                    st.plotly_chart(fig_users, use_container_width=True)
                
                with col2:
                    # Message volume trend
                    fig_messages = px.bar(
                        weekly_data,
                        x='week_display',
                        y='total_messages',
                        title='Weekly Message Volume',
                        labels={'week_display': 'Week Starting', 'total_messages': 'Total Messages'},
                        color='total_messages',
                        color_continuous_scale='Blues'
                    )
                    
                    fig_messages.update_traces(
                        hovertemplate='%{y}<extra></extra>'  # Show only value; <extra></extra> removes trace name box
                    )
                    
                    fig_messages.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='white'),
                        showlegend=False,
                        height=350,
                        xaxis=dict(
                            tickangle=-45,
                            tickmode='auto',
                            nticks=10
                        )
                    )
                    
                    st.plotly_chart(fig_messages, use_container_width=True)
                
                # Weekly summary stats
                st.markdown("**Weekly Summary:**")
                col_a, col_b, col_c = st.columns(3)
                
                with col_a:
                    avg_weekly_users = weekly_data['active_users'].mean()
                    st.metric("Avg Weekly Active Users", f"{avg_weekly_users:,.0f}")
                
                with col_b:
                    avg_weekly_messages = weekly_data['total_messages'].mean()
                    st.metric("Avg Weekly Messages", f"{avg_weekly_messages:,.0f}")
                
                with col_c:
                    # Growth rate (last week vs average)
                    if len(weekly_data) > 1:
                        last_week_users = weekly_data.iloc[-1]['active_users']
                        growth = ((last_week_users - avg_weekly_users) / avg_weekly_users * 100) if avg_weekly_users > 0 else 0
                        st.metric(
                            "Latest Week vs Avg",
                            f"{last_week_users:,.0f}",
                            delta=f"{growth:+.1f}%"
                        )
            
            st.divider()
            
            # Feature Adoption Timeline
            st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">🚀 Feature Adoption Timeline</h3>', unsafe_allow_html=True)
            
            adoption_timeline = get_feature_adoption_timeline(openai_data)
            
            if not adoption_timeline.empty:
                # Count new adopters by feature over time
                adoption_timeline['month'] = pd.to_datetime(adoption_timeline['first_used_date']).dt.to_period('M').astype(str)
                
                monthly_adoption = adoption_timeline.groupby(['month', 'feature']).size().reset_index(name='new_users')
                
                # Create stacked area chart
                fig_adoption = px.area(
                    monthly_adoption,
                    x='month',
                    y='new_users',
                    color='feature',
                    title='Feature Adoption Over Time (New Users per Month)',
                    labels={'month': 'Month', 'new_users': 'New Users Adopting Feature', 'feature': 'Feature'}
                )
                
                fig_adoption.update_layout(
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='white'),
                    hovermode='x unified',
                    height=400
                )
                
                st.plotly_chart(fig_adoption, use_container_width=True)
                
                # Feature adoption summary
                st.markdown("**Feature Adoption Summary:**")
                
                feature_adoption_summary = adoption_timeline.groupby('feature').agg({
                    'user_id': 'count',
                    'first_used_date': 'min'
                }).reset_index()
                feature_adoption_summary.columns = ['Feature', 'Total Adopters', 'First Adoption Date']
                feature_adoption_summary = feature_adoption_summary.sort_values('Total Adopters', ascending=False)
                
                st.dataframe(
                    feature_adoption_summary,
                    use_container_width=True,
                    hide_index=True
                )
            
            st.divider()
            
            # Export Options
            st.markdown('<h3 style="color: var(--text-primary); margin-top: 1.5rem; margin-bottom: 1rem;">📥 OpenAI Data Export</h3>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                # CSV export
                csv_data = openai_data.to_csv(index=False)
                st.download_button(
                    label="📄 Download OpenAI Data (CSV)",
                    data=csv_data,
                    file_name=f"openai_data_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    help="Download all OpenAI ChatGPT usage data as CSV"
                )
            
            with col2:
                # Excel export with summary
                try:
                    from io import BytesIO
                    from openpyxl import Workbook
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    output = BytesIO()
                    wb = Workbook()
                    
                    # Remove default sheet
                    wb.remove(wb.active)
                    
                    # Add data sheet
                    ws_data = wb.create_sheet("OpenAI Data")
                    for r in dataframe_to_rows(openai_data, index=False, header=True):
                        ws_data.append(r)
                    
                    # Add summary sheet
                    ws_summary = wb.create_sheet("Summary", 0)
                    ws_summary.append(["Metric", "Value"])
                    ws_summary.append(["Total Active Users", total_active])
                    ws_summary.append(["Daily Unique Active Users", f"{duau:.0f}"])
                    ws_summary.append(["Total Messages", total_messages])
                    ws_summary.append(["Avg Messages per User", f"{avg_per_user:.0f}"])
                    ws_summary.append(["Avg Days Active per Month", f"{avg_days:.1f}"])
                    
                    # Add message breakdown sheet
                    ws_breakdown = wb.create_sheet("Message Breakdown")
                    for r in dataframe_to_rows(message_breakdown, index=False, header=True):
                        ws_breakdown.append(r)
                    
                    # Add user tiers sheet
                    if not user_tiers.empty:
                        ws_tiers = wb.create_sheet("User Tiers")
                        for r in dataframe_to_rows(user_tiers, index=False, header=True):
                            ws_tiers.append(r)
                    
                    wb.save(output)
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        label="📊 Download OpenAI Report (Excel)",
                        data=excel_data,
                        file_name=f"openai_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="Download comprehensive Excel report with multiple sheets"
                    )
                except ImportError:
                    st.info("Install openpyxl to enable Excel export: pip install openpyxl")
    
    # TAB 3: Power Users
    with tab3:
        st.header("⭐ Power Users & Champions")
        
        # Help text
        st.markdown("""
        <div class="help-tooltip">
            💡 <strong>Power Users</strong> are defined as the top 5% of users by total usage.
            These elite users demonstrate exceptional engagement and are ideal candidates for feedback, beta testing, and advocacy programs.
        </div>
        """, unsafe_allow_html=True)
        
        # Add filtering controls for Power Users
        st.markdown("**🎯 Customize Power User Threshold**")
        filter_col1, filter_col2 = st.columns([1, 1])
        
        with filter_col1:
            # Percentile threshold for power users
            power_user_percentile = st.slider(
                "Top % of users to include",
                min_value=1,
                max_value=25,
                value=5,
                step=1,
                help="Adjust the percentage threshold for identifying power users"
            )
        
        with filter_col2:
            # Minimum message threshold
            min_messages = st.number_input(
                "Min. Total Messages",
                min_value=0,
                max_value=int(data.groupby('user_id')['usage_count'].sum().max()) if not data.empty else 1000,
                value=0,
                help="Only show users with at least this many total messages"
            )
        
        power_users = calculate_power_users(data, threshold_percentile=(100 - power_user_percentile))
        
        # Apply minimum message filter
        if min_messages > 0 and not power_users.empty:
            power_users = power_users[power_users['usage_count'] >= min_messages]
        
        if not power_users.empty:
            # Enhanced metrics row
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric(
                    "Total Power Users", 
                    len(power_users),
                    help="Number of users identified as power users"
                )
            
            with col2:
                pct = (len(power_users) / max(data['user_id'].nunique(), 1)) * 100
                st.metric(
                    "% of Active Users", 
                    f"{pct:.1f}%",
                    help="Percentage of all users who are power users"
                )
            
            with col3:
                power_usage = power_users['usage_count'].sum()
                total_usage_pct = (power_usage / data['usage_count'].sum()) * 100
                st.metric(
                    "Power User Usage", 
                    f"{power_usage:,}",
                    delta=f"{total_usage_pct:.1f}% of total",
                    help="Total usage by power users"
                )
            
            st.divider()
            st.subheader("🏆 Power User Directory")
            st.caption("These users are ideal for feedback, beta testing, and advocacy programs")
            
            # Enhanced table display with better formatting
            for idx, row in power_users.head(20).iterrows():
                # Get message breakdown for this user
                breakdown = get_user_message_breakdown(data, row['email'])
                
                # Calculate totals
                openai_total = sum(breakdown['openai'].values())
                blueflame_total = sum(breakdown['blueflame'].values())
                total_messages = openai_total + blueflame_total
                
                # Create a card-like container for each power user
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #f8f9fa 0%, #ffffff 100%); 
                            padding: 1rem; border-radius: 0.5rem; margin: 0.5rem 0; 
                            border-left: 4px solid #667eea;">
                """, unsafe_allow_html=True)
                
                col1, col2, col3 = st.columns([3, 4, 3])
                
                with col1:
                    # Handle NULL/empty user_name
                    display_name = row['user_name'] if pd.notna(row['user_name']) and row['user_name'] else 'Unknown User'
                    st.write(f"**{display_name}**")
                    # Handle NULL/empty email
                    display_email = row['email'] if pd.notna(row['email']) and row['email'] else 'No email'
                    st.caption(display_email)
                    # Handle NULL/empty department
                    display_dept = row['department'] if pd.notna(row['department']) and row['department'] else 'Unknown'
                    st.caption(f"🏢 {display_dept}")
                
                with col2:
                    st.write("**Message Breakdown:**")
                    
                    # OpenAI Messages Section
                    if openai_total > 0:
                        st.caption(f"**OpenAI Data:** {openai_total:,} messages")
                        if breakdown['openai']['ChatGPT Messages'] > 0:
                            st.caption(f"  💬 ChatGPT: {breakdown['openai']['ChatGPT Messages']:,}")
                        if breakdown['openai']['GPT Messages'] > 0:
                            st.caption(f"  🤖 GPTs: {breakdown['openai']['GPT Messages']:,}")
                        if breakdown['openai']['Tool Messages'] > 0:
                            st.caption(f"  🔧 Tools: {breakdown['openai']['Tool Messages']:,}")
                        if breakdown['openai']['Project Messages'] > 0:
                            st.caption(f"  📁 Projects: {breakdown['openai']['Project Messages']:,}")
                    
                    # BlueFlame Messages Section
                    if blueflame_total > 0:
                        st.caption(f"**BlueFlame Data:** {blueflame_total:,} messages")
                
                with col3:
                    st.write(f"**Total: {total_messages:,}**")
                    # Handle NULL/empty tool_source
                    display_tools = row['tool_source'] if pd.notna(row['tool_source']) and row['tool_source'] else 'Unknown'
                    st.markdown(f'<span class="power-user-badge">{display_tools}</span>', 
                              unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
        else:
            # Enhanced empty state for power users
            st.markdown("""
            <div class="empty-state">
                <div class="empty-state-icon">⭐</div>
                <div class="empty-state-title">No Power Users Yet</div>
                <div class="empty-state-text">
                    Power users will appear here once you have sufficient usage data.<br>
                    Upload more monthly reports to identify your most active users.
                </div>
            </div>
            """, unsafe_allow_html=True)
    
    # TAB 4: Message Type Analytics
    with tab4:
        st.header("📈 Message Type Analytics")
        
        st.markdown("""
        <div class="help-tooltip">
            💡 <strong>Feature Adoption Tracking:</strong> Analyze which AI features are most popular across your organization.
            Use this to identify training opportunities, measure feature adoption, and understand usage patterns.
        </div>
        """, unsafe_allow_html=True)
        
        if data.empty:
            st.info("📊 Upload usage data to see message type analytics.")
        else:
            # Organization-wide feature breakdown
            st.markdown("### 🌐 Organization-Wide Feature Usage")
            
            org_breakdown = get_organization_message_breakdown(data)
            
            if org_breakdown:
                col1, col2 = st.columns([3, 2])
                
                with col1:
                    # Create stacked bar chart showing feature adoption over time
                    st.markdown("**Feature Adoption Trends**")
                    
                    # Get time-series data for each feature
                    time_data = data.copy()
                    time_data['date'] = pd.to_datetime(time_data['date'], errors='coerce')
                    time_data = time_data.dropna(subset=['date'])
                    
                    if not time_data.empty:
                        # Group by month and feature
                        time_data['month'] = time_data['date'].dt.to_period('M').astype(str)
                        feature_trends = time_data.groupby(['month', 'feature_used'])['usage_count'].sum().reset_index()
                        
                        # Create line chart for trends
                        fig_trends = px.line(
                            feature_trends,
                            x='month',
                            y='usage_count',
                            color='feature_used',
                            title='Feature Usage Trends Over Time',
                            labels={'usage_count': 'Message Count', 'month': 'Month', 'feature_used': 'Feature'},
                            markers=True
                        )
                        
                        fig_trends.update_layout(
                            plot_bgcolor='rgba(0,0,0,0)',
                            paper_bgcolor='rgba(0,0,0,0)',
                            font=dict(color='white'),
                            height=400,
                            xaxis=dict(title='Month'),
                            yaxis=dict(title='Messages'),
                            legend=dict(
                                orientation='v',
                                yanchor='top',
                                y=1,
                                xanchor='left',
                                x=1.02
                            )
                        )
                        
                        st.plotly_chart(fig_trends, use_container_width=True)
                    else:
                        st.info("Time-series data not available")
                
                with col2:
                    # Feature usage distribution
                    st.markdown("**Current Distribution**")
                    
                    breakdown_df = pd.DataFrame([
                        {'Feature': feature, 'Messages': count} 
                        for feature, count in org_breakdown.items()
                    ])
                    
                    fig_pie = px.pie(
                        breakdown_df,
                        values='Messages',
                        names='Feature',
                        hole=0.5,
                        color_discrete_sequence=px.colors.qualitative.Pastel
                    )
                    
                    fig_pie.update_traces(
                        textposition='auto',
                        textinfo='percent+label',
                        hovertemplate='<b>%{label}</b><br>Messages: %{value:,}<br>Percentage: %{percent}<extra></extra>'
                    )
                    
                    fig_pie.update_layout(
                        showlegend=False,
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='white'),
                        height=400
                    )
                    
                    st.plotly_chart(fig_pie, use_container_width=True)
                
                st.divider()
                
                # Department-level feature comparison
                st.markdown("### 🏢 Feature Usage by Department")
                st.caption("Compare which features different departments are using")
                
                # Create heatmap of department vs feature usage
                dept_feature_data = data.groupby(['department', 'feature_used'])['usage_count'].sum().reset_index()
                dept_feature_pivot = dept_feature_data.pivot(index='department', columns='feature_used', values='usage_count').fillna(0)
                
                if not dept_feature_pivot.empty:
                    fig_heatmap = px.imshow(
                        dept_feature_pivot,
                        labels=dict(x='Feature Type', y='Department', color='Messages'),
                        aspect='auto',
                        color_continuous_scale='Blues',
                        title='Department vs Feature Usage Heatmap'
                    )
                    
                    fig_heatmap.update_layout(
                        plot_bgcolor='rgba(0,0,0,0)',
                        paper_bgcolor='rgba(0,0,0,0)',
                        font=dict(color='white'),
                        height=max(400, len(dept_feature_pivot) * 30)
                    )
                    
                    fig_heatmap.update_xaxes(side='bottom')
                    
                    st.plotly_chart(fig_heatmap, use_container_width=True)
                    
                    # Add insights
                    st.markdown("### 💡 Feature Adoption Insights")
                    
                    insight_col1, insight_col2, insight_col3 = st.columns(3)
                    
                    with insight_col1:
                        # Most popular feature
                        most_popular_feature = max(org_breakdown.items(), key=lambda x: x[1])
                        st.markdown(f"""
                        <div class="metric-card">
                            <h4>Most Popular Feature</h4>
                            <div class="metric-value">{most_popular_feature[0]}</div>
                            <div class="metric-label">{most_popular_feature[1]:,} messages</div>
                            <div class="metric-label">{(most_popular_feature[1] / sum(org_breakdown.values()) * 100):.1f}% of total</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with insight_col2:
                        # Feature diversity score (number of features being used)
                        active_features = len([f for f, c in org_breakdown.items() if c > 0])
                        all_possible_features = len(get_all_message_types(data))
                        
                        st.markdown(f"""
                        <div class="metric-card">
                            <h4>Feature Adoption Rate</h4>
                            <div class="metric-value">{active_features} / {all_possible_features}</div>
                            <div class="metric-label">Features being used</div>
                            <div class="metric-label">{(active_features / max(all_possible_features, 1) * 100):.0f}% adoption</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with insight_col3:
                        # Department with highest feature diversity
                        dept_diversity = dept_feature_pivot.apply(lambda row: (row > 0).sum(), axis=1)
                        most_diverse_dept = dept_diversity.idxmax()
                        
                        st.markdown(f"""
                        <div class="metric-card">
                            <h4>Most Diverse Usage</h4>
                            <div class="metric-value">{most_diverse_dept}</div>
                            <div class="metric-label">Uses {dept_diversity[most_diverse_dept]} features</div>
                            <div class="metric-label">Highest feature variety</div>
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.info("Not enough data for department comparison")
            else:
                st.info("No feature usage data available")
    
    # TAB ROI: ROI Analytics
    with tab_roi:
        st.header("💎 ROI Analytics")
        
        st.markdown("""
        <div class="help-tooltip">
            💡 <strong>Beyond Traditional Metrics:</strong> This tab provides novel ROI analytics that connect AI usage 
            to business value. Explore time savings, strategic impact, adoption trajectories, and power user dynamics 
            to understand the true return on your AI investment.
        </div>
        """, unsafe_allow_html=True)
        
        if data.empty:
            st.info("📊 Upload usage data to see ROI analytics.")
        else:
            # ========================================================================
            # VISUALIZATION 1: TIME SAVINGS ANALYSIS
            # ========================================================================
            # Methodology: Estimate time saved based on feature usage with industry benchmarks
            # - ChatGPT Messages: avg 10 minutes saved per complex query (research, writing, analysis)
            # - Tool Messages: avg 30 minutes saved per task (code generation, data analysis)
            # - Project Messages: avg 15 minutes saved per collaborative task
            # - GPT Messages: avg 20 minutes saved per custom GPT interaction
            # - BlueFlame Messages: avg 12 minutes saved per enterprise query
            # 
            # These estimates are conservative and based on typical task completion times.
            # Organizations should calibrate these values based on their specific use cases.
            # ========================================================================
            
            st.markdown("### ⏱️ Time Savings Analysis")
            st.caption("Estimated productivity gains from AI tool usage based on industry benchmarks")
            
            # Define time savings per message type (in minutes)
            time_savings_map = {
                'ChatGPT Messages': 10,      # General AI assistance
                'Tool Messages': 30,          # Code Interpreter, advanced tools
                'Project Messages': 15,       # Collaborative workspace usage
                'GPT Messages': 20,           # Custom GPT interactions
                'BlueFlame Messages': 12      # Enterprise-specific queries
            }
            
            # Calculate time savings by feature and department
            time_data = data.copy()
            time_data['minutes_saved'] = time_data.apply(
                lambda row: row['usage_count'] * time_savings_map.get(row['feature_used'], 10),
                axis=1
            )
            time_data['hours_saved'] = time_data['minutes_saved'] / 60
            
            # Overall time savings metrics
            total_hours_saved = time_data['hours_saved'].sum()
            total_days_saved = total_hours_saved / 8  # 8-hour workday
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "Total Hours Saved",
                    f"{total_hours_saved:,.0f}",
                    help="Estimated cumulative time savings across all AI interactions"
                )
            
            with col2:
                st.metric(
                    "Work Days Saved",
                    f"{total_days_saved:,.0f}",
                    help="Equivalent full work days saved (8-hour days)"
                )
            
            with col3:
                avg_hours_per_user = total_hours_saved / max(data['user_id'].nunique(), 1)
                st.metric(
                    "Avg Hours/User",
                    f"{avg_hours_per_user:,.1f}",
                    help="Average time savings per active user"
                )
            
            with col4:
                # Calculate monthly productivity boost as percentage
                # Using configurable monthly work hours (default: 160 hours = 20 days × 8 hours)
                total_users = data['user_id'].nunique()
                available_hours = total_users * ROI_MONTHLY_WORK_HOURS
                productivity_boost = (total_hours_saved / max(available_hours, 1)) * 100
                st.metric(
                    "Productivity Boost",
                    f"{productivity_boost:,.1f}%",
                    help="Estimated productivity increase based on time saved vs. available work hours"
                )
            
            # Time savings by feature type
            st.markdown("**Time Savings Breakdown by AI Feature**")
            
            savings_by_feature = time_data.groupby('feature_used').agg({
                'hours_saved': 'sum',
                'usage_count': 'sum'
            }).reset_index()
            savings_by_feature = savings_by_feature.sort_values('hours_saved', ascending=False)
            
            fig_time_savings = px.bar(
                savings_by_feature,
                x='feature_used',
                y='hours_saved',
                title='Estimated Time Savings by Feature Type',
                labels={'hours_saved': 'Hours Saved', 'feature_used': 'Feature'},
                color='hours_saved',
                color_continuous_scale='Viridis',
                text='hours_saved'
            )
            
            fig_time_savings.update_traces(
                texttemplate='%{text:,.0f}h',
                textposition='outside',
                hovertemplate='<b>%{x}</b><br>Hours Saved: %{y:,.0f}<br>Messages: %{customdata[0]:,}<extra></extra>',
                customdata=savings_by_feature[['usage_count']].values
            )
            
            fig_time_savings.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white'),
                height=400,
                showlegend=False,
                xaxis_title='AI Feature',
                yaxis_title='Hours Saved'
            )
            
            st.plotly_chart(fig_time_savings, use_container_width=True)
            
            # Methodology note
            with st.expander("📊 Calculation Methodology"):
                st.markdown(f"""
                **Time Savings Estimates:**
                
                These estimates are based on industry research and conservative assumptions about time saved per AI interaction:
                
                {chr(10).join([f'- **{feature}**: {minutes} minutes saved per message' for feature, minutes in time_savings_map.items()])}
                
                **Total Calculation:**
                - Messages: {data['usage_count'].sum():,}
                - Estimated Hours Saved: {total_hours_saved:,.0f}
                - Equivalent Work Days: {total_days_saved:,.0f}
                
                **Note:** These are conservative estimates. Actual time savings may vary based on:
                - Task complexity and context
                - User proficiency with AI tools
                - Availability of alternative solutions
                - Quality of AI-generated outputs
                
                Organizations should calibrate these values based on internal benchmarking and user surveys.
                """)
            
            st.divider()
            
            # ========================================================================
            # VISUALIZATION 2: ADOPTION TRAJECTORY & VALUE TRENDS
            # ========================================================================
            # Methodology: Track cumulative business value over time using composite metrics
            # - Cumulative messages as proxy for AI integration depth
            # - Active user growth rate as adoption velocity indicator
            # - Feature diversity score as capability maturity metric
            # - Combined into a "Value Trajectory Score" normalized 0-100
            # 
            # This reveals whether the organization is in:
            # - Early adoption (growing but low volume)
            # - Scaling phase (rapid growth)
            # - Mature adoption (high volume, stable growth)
            # ========================================================================
            
            st.markdown("### 📈 Adoption Trajectory & Value Trends")
            st.caption("Track the evolution of AI integration and business value over time")
            
            # Prepare time-series data
            ts_data = data.copy()
            ts_data['date'] = pd.to_datetime(ts_data['date'], errors='coerce')
            ts_data = ts_data.dropna(subset=['date'])
            
            if not ts_data.empty:
                # Group by month for trend analysis
                ts_data['month'] = ts_data['date'].dt.to_period('M').astype(str)
                
                # Calculate monthly metrics
                monthly_stats = []
                
                for month in sorted(ts_data['month'].unique()):
                    month_data = ts_data[ts_data['month'] == month]
                    
                    # Core metrics
                    active_users = month_data['user_id'].nunique()
                    total_messages = month_data['usage_count'].sum()
                    unique_features = month_data['feature_used'].nunique()
                    
                    # Time savings for this month
                    month_data['hours_saved'] = month_data.apply(
                        lambda row: row['usage_count'] * time_savings_map.get(row['feature_used'], 10) / 60,
                        axis=1
                    )
                    hours_saved = month_data['hours_saved'].sum()
                    
                    # Composite Value Score (normalized 0-100)
                    # Components:
                    # - User adoption (30%): Active users relative to max observed
                    # - Usage intensity (30%): Messages per user
                    # - Feature diversity (20%): Features used / total available
                    # - Time value (20%): Hours saved
                    
                    monthly_stats.append({
                        'month': month,
                        'active_users': active_users,
                        'total_messages': total_messages,
                        'messages_per_user': total_messages / max(active_users, 1),
                        'unique_features': unique_features,
                        'hours_saved': hours_saved,
                        'value_delivered': hours_saved * ROI_HOURLY_VALUE  # Configurable hourly value
                    })
                
                monthly_df = pd.DataFrame(monthly_stats)
                
                # Calculate composite value score
                max_users = monthly_df['active_users'].max()
                max_msgs_per_user = monthly_df['messages_per_user'].max()
                max_features = monthly_df['unique_features'].max()
                max_hours = monthly_df['hours_saved'].max()
                
                monthly_df['value_score'] = (
                    (monthly_df['active_users'] / max(max_users, 1) * 30) +
                    (monthly_df['messages_per_user'] / max(max_msgs_per_user, 1) * 30) +
                    (monthly_df['unique_features'] / max(max_features, 1) * 20) +
                    (monthly_df['hours_saved'] / max(max_hours, 1) * 20)
                )
                
                # Create dual-axis chart: cumulative messages & value score
                fig_trajectory = make_subplots(
                    specs=[[{"secondary_y": True}]]
                )
                
                # Cumulative messages (primary axis)
                monthly_df['cumulative_messages'] = monthly_df['total_messages'].cumsum()
                
                fig_trajectory.add_trace(
                    go.Scatter(
                        x=monthly_df['month'],
                        y=monthly_df['cumulative_messages'],
                        name='Cumulative Messages',
                        mode='lines+markers',
                        line=dict(color='#00D9FF', width=3),
                        fill='tozeroy',
                        fillcolor='rgba(0, 217, 255, 0.1)',
                        hovertemplate='<b>%{x}</b><br>Cumulative Messages: %{y:,}<extra></extra>'
                    ),
                    secondary_y=False
                )
                
                # Value score (secondary axis)
                fig_trajectory.add_trace(
                    go.Scatter(
                        x=monthly_df['month'],
                        y=monthly_df['value_score'],
                        name='Value Trajectory Score',
                        mode='lines+markers',
                        line=dict(color='#FFD700', width=3, dash='dot'),
                        marker=dict(size=10),
                        hovertemplate='<b>%{x}</b><br>Value Score: %{y:.1f}/100<extra></extra>'
                    ),
                    secondary_y=True
                )
                
                fig_trajectory.update_layout(
                    title='AI Adoption & Value Creation Over Time',
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(color='white'),
                    height=450,
                    hovermode='x unified',
                    legend=dict(
                        orientation='h',
                        yanchor='bottom',
                        y=1.02,
                        xanchor='right',
                        x=1
                    )
                )
                
                fig_trajectory.update_xaxes(title_text='Month')
                fig_trajectory.update_yaxes(title_text='Cumulative Messages', secondary_y=False, color='#00D9FF')
                fig_trajectory.update_yaxes(title_text='Value Score (0-100)', secondary_y=True, color='#FFD700')
                
                st.plotly_chart(fig_trajectory, use_container_width=True)
                
                # Growth insights
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if len(monthly_df) >= 2:
                        first_month_users = monthly_df.iloc[0]['active_users']
                        last_month_users = monthly_df.iloc[-1]['active_users']
                        user_growth = ((last_month_users - first_month_users) / max(first_month_users, 1)) * 100
                        st.metric(
                            "User Growth",
                            f"{user_growth:+.1f}%",
                            help="Change in active users from first to last month"
                        )
                
                with col2:
                    total_value_delivered = monthly_df['value_delivered'].sum()
                    st.metric(
                        "Est. Value Created",
                        f"${total_value_delivered:,.0f}",
                        help=f"Estimated business value based on time saved (${ROI_HOURLY_VALUE}/hour)"
                    )
                
                with col3:
                    current_value_score = monthly_df.iloc[-1]['value_score'] if len(monthly_df) > 0 else 0
                    st.metric(
                        "Current Value Score",
                        f"{current_value_score:.1f}/100",
                        help="Composite metric: user adoption + usage intensity + feature diversity + time value"
                    )
                
                # Methodology note
                with st.expander("📊 Value Score Methodology"):
                    st.markdown("""
                    **Composite Value Trajectory Score (0-100):**
                    
                    This novel metric combines multiple dimensions of AI ROI:
                    
                    1. **User Adoption (30%)**: Active users relative to peak month
                    2. **Usage Intensity (30%)**: Messages per user (engagement depth)
                    3. **Feature Diversity (20%)**: Number of different AI features being used
                    4. **Time Value (20%)**: Hours saved through AI assistance
                    
                    **Score Interpretation:**
                    - **0-30**: Early adoption phase - building awareness
                    - **31-60**: Scaling phase - growing engagement
                    - **61-85**: Mature adoption - organization-wide integration
                    - **86-100**: Advanced optimization - maximizing value
                    
                    **Why This Matters:**
                    This score goes beyond simple usage counts to measure true business impact. 
                    A rising trajectory indicates successful AI integration delivering increasing value.
                    """)
            else:
                st.info("Time-series data not available for trajectory analysis")
            
            st.divider()
            
            # ========================================================================
            # VISUALIZATION 3: POWER USER & DEPARTMENT IMPACT MATRIX
            # ========================================================================
            # Methodology: Capability quadrant analysis plotting departments/users on two axes:
            # - X-axis: Usage Volume (messages sent) - represents engagement level
            # - Y-axis: Feature Diversity (unique features used) - represents capability breadth
            # 
            # Quadrants:
            # - High Volume, High Diversity: "Strategic Leaders" - driving innovation
            # - High Volume, Low Diversity: "Focused Power Users" - deep expertise in specific tools
            # - Low Volume, High Diversity: "Explorers" - broad experimentation, growth potential
            # - Low Volume, Low Diversity: "Emerging Users" - need training/enablement
            # 
            # Bubble size represents estimated business value (time saved * $50/hour)
            # ========================================================================
            
            st.markdown("### 🎯 Department Impact Matrix (Capability Quadrant)")
            st.caption("Identify strategic leaders, power users, explorers, and emerging adopters")
            
            # Calculate department-level metrics
            dept_metrics = []
            
            for dept in data['department'].unique():
                if pd.isna(dept) or dept == '':
                    dept = 'Unknown'
                
                dept_data = data[data['department'] == dept]
                
                total_messages = dept_data['usage_count'].sum()
                unique_features = dept_data['feature_used'].nunique()
                active_users = dept_data['user_id'].nunique()
                
                # Calculate time savings
                dept_data_copy = dept_data.copy()
                dept_data_copy['hours_saved'] = dept_data_copy.apply(
                    lambda row: row['usage_count'] * time_savings_map.get(row['feature_used'], 10) / 60,
                    axis=1
                )
                hours_saved = dept_data_copy['hours_saved'].sum()
                value_estimate = hours_saved * ROI_HOURLY_VALUE  # Configurable hourly value
                
                dept_metrics.append({
                    'department': dept,
                    'total_messages': total_messages,
                    'unique_features': unique_features,
                    'active_users': active_users,
                    'messages_per_user': total_messages / max(active_users, 1),
                    'hours_saved': hours_saved,
                    'value_estimate': value_estimate
                })
            
            dept_df = pd.DataFrame(dept_metrics)
            
            # Create quadrant scatter plot
            fig_quadrant = px.scatter(
                dept_df,
                x='total_messages',
                y='unique_features',
                size='value_estimate',
                color='messages_per_user',
                hover_name='department',
                hover_data={
                    'total_messages': ':,',
                    'unique_features': True,
                    'active_users': True,
                    'hours_saved': ':.0f',
                    'value_estimate': ':$,.0f',
                    'messages_per_user': ':.1f'
                },
                labels={
                    'total_messages': 'Usage Volume (Messages)',
                    'unique_features': 'Feature Diversity (Unique Features)',
                    'messages_per_user': 'Avg Messages/User',
                    'value_estimate': 'Est. Value'
                },
                title='Department Capability Quadrant Analysis',
                color_continuous_scale='Viridis',
                size_max=60
            )
            
            # Add quadrant lines (median splits)
            median_messages = dept_df['total_messages'].median()
            median_features = dept_df['unique_features'].median()
            
            fig_quadrant.add_hline(
                y=median_features,
                line_dash='dash',
                line_color='rgba(255,255,255,0.3)',
                annotation_text='Median Feature Diversity',
                annotation_position='right'
            )
            
            fig_quadrant.add_vline(
                x=median_messages,
                line_dash='dash',
                line_color='rgba(255,255,255,0.3)',
                annotation_text='Median Usage',
                annotation_position='top'
            )
            
            # Add quadrant labels
            max_messages = dept_df['total_messages'].max()
            max_features = dept_df['unique_features'].max()
            
            annotations = [
                dict(x=median_messages * 1.5, y=max_features * 0.95, text='<b>Strategic Leaders</b><br>High usage + High diversity',
                     showarrow=False, font=dict(size=10, color='rgba(255,255,255,0.7)'), align='center'),
                dict(x=median_messages * 1.5, y=median_features * 0.3, text='<b>Focused Power Users</b><br>High usage + Specialized',
                     showarrow=False, font=dict(size=10, color='rgba(255,255,255,0.7)'), align='center'),
                dict(x=median_messages * 0.3, y=max_features * 0.95, text='<b>Explorers</b><br>Experimenting broadly',
                     showarrow=False, font=dict(size=10, color='rgba(255,255,255,0.7)'), align='center'),
                dict(x=median_messages * 0.3, y=median_features * 0.3, text='<b>Emerging Users</b><br>Growth potential',
                     showarrow=False, font=dict(size=10, color='rgba(255,255,255,0.7)'), align='center'),
            ]
            
            fig_quadrant.update_layout(
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                font=dict(color='white'),
                height=500,
                annotations=annotations,
                xaxis_title='Usage Volume (Total Messages)',
                yaxis_title='Feature Diversity (Unique Features Used)'
            )
            
            st.plotly_chart(fig_quadrant, use_container_width=True)
            
            # Identify top performers in each quadrant
            st.markdown("**Department Categorization:**")
            
            quad_col1, quad_col2, quad_col3, quad_col4 = st.columns(4)
            
            with quad_col1:
                strategic_leaders = dept_df[
                    (dept_df['total_messages'] >= median_messages) & 
                    (dept_df['unique_features'] >= median_features)
                ].sort_values('value_estimate', ascending=False)
                
                st.markdown("**🌟 Strategic Leaders**")
                if not strategic_leaders.empty:
                    for _, dept in strategic_leaders.head(3).iterrows():
                        st.markdown(f"- {dept['department']}")
                        st.caption(f"  {dept['total_messages']:,} msgs, {dept['unique_features']} features")
                else:
                    st.caption("None identified")
            
            with quad_col2:
                power_users = dept_df[
                    (dept_df['total_messages'] >= median_messages) & 
                    (dept_df['unique_features'] < median_features)
                ].sort_values('total_messages', ascending=False)
                
                st.markdown("**💪 Focused Power Users**")
                if not power_users.empty:
                    for _, dept in power_users.head(3).iterrows():
                        st.markdown(f"- {dept['department']}")
                        st.caption(f"  {dept['total_messages']:,} msgs, {dept['unique_features']} features")
                else:
                    st.caption("None identified")
            
            with quad_col3:
                explorers = dept_df[
                    (dept_df['total_messages'] < median_messages) & 
                    (dept_df['unique_features'] >= median_features)
                ].sort_values('unique_features', ascending=False)
                
                st.markdown("**🔍 Explorers**")
                if not explorers.empty:
                    for _, dept in explorers.head(3).iterrows():
                        st.markdown(f"- {dept['department']}")
                        st.caption(f"  {dept['total_messages']:,} msgs, {dept['unique_features']} features")
                else:
                    st.caption("None identified")
            
            with quad_col4:
                emerging = dept_df[
                    (dept_df['total_messages'] < median_messages) & 
                    (dept_df['unique_features'] < median_features)
                ].sort_values('active_users', ascending=False)
                
                st.markdown("**🌱 Emerging Users**")
                if not emerging.empty:
                    for _, dept in emerging.head(3).iterrows():
                        st.markdown(f"- {dept['department']}")
                        st.caption(f"  {dept['active_users']} users, {dept['total_messages']:,} msgs")
                else:
                    st.caption("None identified")
            
            # Methodology note
            with st.expander("📊 Quadrant Analysis Methodology"):
                st.markdown("""
                **Capability Quadrant Framework:**
                
                This analysis maps departments on two critical dimensions:
                
                **X-Axis: Usage Volume**
                - Measures total messages sent by department
                - Indicates level of AI tool engagement and dependency
                - Higher values suggest deeper integration into workflows
                
                **Y-Axis: Feature Diversity**
                - Counts unique AI features being utilized
                - Represents capability breadth and tool sophistication
                - Higher values indicate mature, multi-faceted AI adoption
                
                **Bubble Size: Estimated Business Value**
                - Based on time saved (hours) × ${ROI_HOURLY_VALUE}/hour (configurable)
                - Larger bubbles = greater estimated ROI
                - Helps prioritize high-impact departments
                
                **Quadrant Interpretation:**
                
                1. **Strategic Leaders** (High Usage + High Diversity)
                   - Driving AI innovation across the organization
                   - Should be showcased as success stories
                   - Consider for peer mentoring programs
                
                2. **Focused Power Users** (High Usage + Low Diversity)
                   - Deep expertise in specific AI tools
                   - Opportunity: Introduce adjacent features for expansion
                   - May benefit from advanced training on other capabilities
                
                3. **Explorers** (Low Usage + High Diversity)
                   - Experimenting broadly with different features
                   - High growth potential with proper support
                   - Target for success stories to boost confidence
                
                4. **Emerging Users** (Low Usage + Low Diversity)
                   - Early in AI adoption journey
                   - Require training, awareness, and enablement
                   - Focus on quick wins to build momentum
                
                **Strategic Actions:**
                - Move Emerging → Explorers: Training & awareness campaigns
                - Move Explorers → Strategic Leaders: Success showcases & support
                - Move Power Users → Strategic Leaders: Feature education
                - Maintain Strategic Leaders: Advanced workshops & innovation programs
                """)
            
            # Summary insights
            st.divider()
            st.markdown("### 🎯 Strategic ROI Insights")
            
            insight1, insight2, insight3 = st.columns(3)
            
            with insight1:
                st.markdown("""
                <div class="metric-card">
                    <h4>💰 Total Estimated Value</h4>
                    <div class="metric-value">${:,.0f}</div>
                    <div class="metric-label">Based on time saved</div>
                </div>
                """.format(dept_df['value_estimate'].sum()), unsafe_allow_html=True)
            
            with insight2:
                top_dept = dept_df.loc[dept_df['value_estimate'].idxmax()] if not dept_df.empty else None
                if top_dept is not None:
                    st.markdown(f"""
                    <div class="metric-card">
                        <h4>🏆 Highest Impact Department</h4>
                        <div class="metric-value">{top_dept['department']}</div>
                        <div class="metric-label">${top_dept['value_estimate']:,.0f} value created</div>
                    </div>
                    """, unsafe_allow_html=True)
            
            with insight3:
                avg_value_per_dept = dept_df['value_estimate'].mean()
                st.markdown(f"""
                <div class="metric-card">
                    <h4>📊 Avg Value per Department</h4>
                    <div class="metric-value">${avg_value_per_dept:,.0f}</div>
                    <div class="metric-label">Opportunity for underperformers</div>
                </div>
                """, unsafe_allow_html=True)
    
    # TAB 5: Department Mapper
    with tab5:
        display_department_mapper()
    
    # TAB 6: Database Management
    with tab6:
        st.header("🔧 Database Management")
        
        st.markdown("""
        <div class="help-tooltip">
            💡 Monitor your database health, manage uploads, and export data for external analysis
        </div>
        """, unsafe_allow_html=True)
        
        db_info = get_database_info()
        
        # Enhanced database stats
        st.markdown('<div class="section-header"><h3>📊 Database Statistics</h3></div>', unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric(
                "Total Records", 
                f"{db_info['total_stats']['total_records']:,}",
                help="Total number of records in database"
            )
        with col2:
            st.metric(
                "Total Users", 
                db_info['total_stats']['total_users'],
                help="Unique users across all data"
            )
        with col3:
            st.metric(
                "Date Range", 
                f"{db_info['total_stats']['total_days']} days",
                help="Total days of data coverage"
            )
        
        st.divider()
        
        # Employee File Management
        st.markdown('<div class="section-header"><h3>👥 Employee Master File</h3></div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="help-tooltip">
            📋 Upload your employee roster to enable department assignment and identify non-employee tool usage
        </div>
        """, unsafe_allow_html=True)
        
        try:
            employee_count = db.get_employee_count()
        except AttributeError:
            # Handle cache error - database object missing methods
            employee_count = 0
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.metric("Employees Loaded", f"{employee_count:,}", help="Total employees in master roster")
        
        with col2:
            if st.button("👁️ View Employees", help="View all loaded employees"):
                st.session_state.show_employees = not st.session_state.get('show_employees', False)
        
        # Show employee table if requested
        if st.session_state.get('show_employees', False):
            try:
                employees_df = db.get_all_employees()
            except AttributeError:
                # Handle cache error - database object missing methods
                employees_df = pd.DataFrame()
            
            if not employees_df.empty:
                st.markdown("**Employee Actions:**")
                st.caption("⚠️ Deleting an employee will remove them from the roster and optionally remove their usage data from all analytics.")
                
                # Add search filter for employees
                search_emp = st.text_input("🔍 Search employees by name or email", "", key="employee_search")
                
                # Filter employees if search is active
                if search_emp:
                    employees_filtered = employees_df[
                        employees_df['first_name'].str.contains(search_emp, case=False, na=False) |
                        employees_df['last_name'].str.contains(search_emp, case=False, na=False) |
                        employees_df['email'].str.contains(search_emp, case=False, na=False)
                    ]
                else:
                    employees_filtered = employees_df
                
                if employees_filtered.empty:
                    st.info(f"No employees match '{search_emp}'")
                else:
                    st.write(f"**Showing {len(employees_filtered)} of {len(employees_df)} employees:**")
                    
                    # Display employees with delete options
                    for idx, row in employees_filtered.iterrows():
                        col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 2, 1])
                        
                        with col1:
                            st.write(f"**{row['first_name']} {row['last_name']}**")
                        
                        with col2:
                            if pd.notna(row['email']) and row['email']:
                                st.write(f"📧 {row['email']}")
                            else:
                                st.write("📧 _No email_")
                        
                        with col3:
                            if pd.notna(row['department']) and row['department']:
                                st.write(f"🏢 {row['department']}")
                            else:
                                st.write("🏢 _Unknown_")
                        
                        with col4:
                            if pd.notna(row['status']) and row['status']:
                                st.write(f"Status: {row['status']}")
                            else:
                                st.write("Status: _Unknown_")
                        
                        with col5:
                            # Delete button with confirmation
                            delete_key = f"delete_emp_{row['employee_id']}"
                            if st.button("🗑️", key=delete_key, help=f"Delete {row['first_name']} {row['last_name']}"):
                                st.session_state[f"confirm_{delete_key}"] = True
                        
                        # Confirmation dialog if delete was clicked
                        confirm_key = f"confirm_delete_emp_{row['employee_id']}"
                        if st.session_state.get(confirm_key, False):
                            st.error(f"⚠️ **Confirm deletion of {row['first_name']} {row['last_name']}**")
                            
                            col_a, col_b, col_c = st.columns(3)
                            
                            with col_a:
                                if st.button("❌ Cancel", key=f"cancel_{delete_key}"):
                                    st.session_state[confirm_key] = False
                                    st.rerun()
                            
                            with col_b:
                                if st.button("🗑️ Delete Employee Only", key=f"confirm_emp_{delete_key}", 
                                           help="Remove from employee roster but keep usage data"):
                                    success, message = db.delete_employee(row['employee_id'])
                                    if success:
                                        st.success(f"✅ {message}")
                                        st.session_state[confirm_key] = False
                                        st.rerun()
                                    else:
                                        st.error(f"❌ {message}")
                            
                            with col_c:
                                if st.button("💣 Delete All Data", key=f"confirm_all_{delete_key}", 
                                           help="Remove employee AND all their usage data from analytics"):
                                    success, message = db.delete_employee_and_usage(row['employee_id'])
                                    if success:
                                        st.success(f"✅ {message}")
                                        st.session_state[confirm_key] = False
                                        st.cache_resource.clear()  # Clear cache to refresh metrics
                                        st.rerun()
                                    else:
                                        st.error(f"❌ {message}")
                        
                        st.divider()
            else:
                st.info("No employees loaded yet")
        
        # Employee file uploader
        employee_file = st.file_uploader(
            "📤 Upload Employee Master File",
            type=['csv', 'xlsx'],
            help="Upload CSV or Excel file with columns: First Name, Last Name, Email, Title, Function (for department), Status",
            key="employee_uploader"
        )
        
        if employee_file is not None:
            try:
                # Read the file
                if employee_file.name.endswith('.csv'):
                    emp_df = pd.read_csv(employee_file, low_memory=False)
                else:
                    emp_df = pd.read_excel(employee_file)
                
                st.write(f"**Preview of {employee_file.name}:**")
                st.dataframe(emp_df.head(5), use_container_width=True)
                
                # Map columns
                st.write("**Column Mapping:**")
                st.info("ℹ️ Email column is optional. If your file doesn't have email addresses, employees will be matched by name.")
                col_map_col1, col_map_col2 = st.columns(2)
                
                with col_map_col1:
                    first_name_col = st.selectbox("First Name Column", emp_df.columns, 
                                                  index=next((i for i, c in enumerate(emp_df.columns) if 'first' in c.lower()), 0))
                    last_name_col = st.selectbox("Last Name Column", emp_df.columns,
                                                index=next((i for i, c in enumerate(emp_df.columns) if 'last' in c.lower()), 1))
                    # Make email optional
                    email_col_options = ['[No Email Column]'] + list(emp_df.columns)
                    email_col_idx = next((i+1 for i, c in enumerate(emp_df.columns) if 'email' in c.lower()), 0)
                    email_col_selection = st.selectbox("Email Column (Optional)", email_col_options, index=email_col_idx)
                
                with col_map_col2:
                    title_col = st.selectbox("Title Column", emp_df.columns,
                                           index=next((i for i, c in enumerate(emp_df.columns) if 'title' in c.lower() or 'position' in c.lower()), 3))
                    dept_col = st.selectbox("Department Column (Function)", emp_df.columns,
                                          index=next((i for i, c in enumerate(emp_df.columns) if 'function' in c.lower() or 'department' in c.lower() or 'dept' in c.lower()), 4))
                    status_col = st.selectbox("Status Column", emp_df.columns,
                                            index=next((i for i, c in enumerate(emp_df.columns) if 'status' in c.lower()), 5))
                
                if st.button("📥 Load Employees", type="primary"):
                    # Create normalized dataframe
                    normalized_emp_df = pd.DataFrame({
                        'first_name': emp_df[first_name_col],
                        'last_name': emp_df[last_name_col],
                        'email': emp_df[email_col_selection] if email_col_selection != '[No Email Column]' else None,
                        'title': emp_df[title_col],
                        'department': emp_df[dept_col],
                        'status': emp_df[status_col]
                    })
                    
                    # Load into database
                    success, message, count = db.load_employees(normalized_emp_df)
                    
                    if success:
                        st.success(f"✅ {message}")
                        st.rerun()
                    else:
                        st.error(f"❌ {message}")
                        
            except Exception as e:
                st.error(f"❌ Error processing employee file: {str(e)}")
        
        # Force reload employee file button
        if employee_count > 0:
            st.markdown("**🔄 Employee File Actions**")
            col_reload1, col_reload2 = st.columns(2)
            
            with col_reload1:
                if st.button("🔄 Force Re-import Employee File", 
                           help="Clear marker and reload the employee CSV file from disk",
                           use_container_width=True):
                    with st.spinner("Reloading employee file..."):
                        success, message = force_reload_employee_file()
                        if success:
                            st.success(message)
                            st.cache_resource.clear()  # Clear cache
                            st.rerun()
                        else:
                            st.error(message)
        
        st.divider()
        
        # Upload history with better styling and file management
        st.markdown('<div class="section-header"><h3>📂 Upload History & File Management</h3></div>', unsafe_allow_html=True)
        
        if db_info['upload_history']:
            upload_df = pd.DataFrame(db_info['upload_history'])
            
            # Display upload history with delete option
            for idx, row in upload_df.iterrows():
                col1, col2, col3, col4 = st.columns([3, 2, 1, 1])
                
                with col1:
                    st.write(f"**📄 {row['filename']}**")
                
                with col2:
                    st.write(f"📅 {row['date_range']}")
                
                with col3:
                    st.write(f"📊 {row['records']:,} records")
                
                with col4:
                    # Delete button for individual file
                    if st.button("🗑️", key=f"delete_file_{idx}", help=f"Delete {row['filename']}"):
                        confirm_key = f"confirm_delete_{idx}"
                        if st.session_state.get(confirm_key, False):
                            # Actually delete the file
                            success = db.delete_by_file(row['filename'])
                            if success:
                                st.success(f"✅ Deleted {row['filename']}")
                                st.session_state[confirm_key] = False
                                st.rerun()
                            else:
                                st.error(f"❌ Error deleting {row['filename']}")
                        else:
                            # Set confirmation flag
                            st.session_state[confirm_key] = True
                            st.warning(f"⚠️ Click again to confirm deletion of {row['filename']}")
                
                # Show confirmation warning if needed
                confirm_key = f"confirm_delete_{idx}"
                if st.session_state.get(confirm_key, False):
                    st.error(f"⚠️ **Confirm deletion:** Click 🗑️ again to permanently delete {row['filename']} ({row['records']:,} records)")
                
                st.divider()
        else:
            st.markdown("""
            <div class="empty-state">
                <div class="empty-state-icon">📁</div>
                <div class="empty-state-title">No Upload History</div>
                <div class="empty-state-text">Upload your first file to see history here</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.divider()
        
        # Database actions with better UX
        st.markdown('<div class="section-header"><h3>⚙️ Database Actions</h3></div>', unsafe_allow_html=True)
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.write("**Export Data**")
            all_data = db.get_all_data()
            if not all_data.empty:
                csv = all_data.to_csv(index=False)
                st.download_button(
                    "📥 Download CSV",
                    csv,
                    f"ai_usage_export_{datetime.now().strftime('%Y%m%d')}.csv",
                    "text/csv",
                    use_container_width=True,
                    help="Download all data as CSV for external analysis"
                )
            else:
                st.button("📥 Download CSV", disabled=True, use_container_width=True)
        
        with col2:
            st.write("**Refresh Data**")
            if st.button("🔄 Refresh Dashboard", type="secondary", use_container_width=True):
                st.cache_resource.clear()
                st.success("Cache cleared!")
                st.rerun()
        
        with col3:
            st.write("**Clear Database**")
            if st.button("🗑️ Clear All Data", type="secondary", use_container_width=True):
                if st.session_state.get('confirm_clear'):
                    db.delete_all_data()
                    st.session_state.confirm_clear = False
                    st.success("✅ Database cleared successfully")
                    st.rerun()
                else:
                    st.session_state.confirm_clear = True
                    st.warning("⚠️ Click again to confirm deletion")
        
        with col4:
            st.write("**Clear & Reset All**")
            if st.button("💣 Clear & Reset", type="secondary", use_container_width=True, 
                        help="Clear database, file tracking, and employee markers"):
                if st.session_state.get('confirm_full_reset'):
                    with st.spinner("Performing complete reset..."):
                        success, message, details = clear_and_reset_all()
                        st.session_state.confirm_full_reset = False
                        
                        if success:
                            st.success(message)
                            st.cache_resource.clear()
                            st.rerun()
                        else:
                            st.warning(message)
                            st.json(details)
                else:
                    st.session_state.confirm_full_reset = True
                    st.error("⚠️ Click again to confirm COMPLETE RESET")
        
        # Warning about clear action
        if st.session_state.get('confirm_clear'):
            st.error("⚠️ **Warning:** This action will permanently delete all data. Click the button again to confirm.")
        
        # Warning about full reset
        if st.session_state.get('confirm_full_reset'):
            st.error("""
            ⚠️ **WARNING: COMPLETE RESET**
            
            This will:
            - Delete ALL database records
            - Remove file tracking history (all files will appear as new)
            - Clear employee file markers (allowing re-import)
            
            Click "Clear & Reset" again to confirm this action.
            """)

def get_database_info():
    """Get database information."""
    all_data = db.get_all_data()
    
    if all_data.empty:
        return {
            'total_stats': {'total_records': 0, 'total_users': 0, 'total_days': 0, 'total_cost': 0.0},
            'upload_history': [],
            'date_coverage': pd.DataFrame()
        }
    
    total_cost = all_data['cost_usd'].sum() if 'cost_usd' in all_data.columns else 0.0
    
    # Calculate total days with robust date handling
    try:
        valid_dates = pd.to_datetime(all_data['date'], errors='coerce').dropna()
        if len(valid_dates) > 0:
            total_days = (valid_dates.max() - valid_dates.min()).days + 1
        else:
            total_days = 0
    except Exception:
        total_days = 0
    
    total_stats = {
        'total_records': len(all_data),
        'total_users': all_data['user_id'].nunique(),
        'total_days': total_days,
        'total_cost': float(total_cost)
    }
    
    upload_history = []
    if 'file_source' in all_data.columns:
        for filename in all_data['file_source'].unique():
            file_data = all_data[all_data['file_source'] == filename]
            upload_history.append({
                'filename': filename,
                'date_range': f"{file_data['date'].min()} to {file_data['date'].max()}",
                'records': len(file_data)
            })
    
    return {
        'total_stats': total_stats,
        'upload_history': upload_history,
        'date_coverage': pd.DataFrame()
    }

if __name__ == "__main__":
    main()